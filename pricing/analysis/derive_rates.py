"""
Verify landed-cost model against pricefeed xlsx base_cost.

Reads base_cost (full landed cost) from S3 pricefeed xlsx files,
reads cost_gbp from RDS, and compares against the current model:

    landed_cost_gbp = cost_gbp * (1 + shipping_rate) + shipping_flat

Usage:
    python -m pricing.analysis.derive_rates [--bucket BUCKET] [--dry-run]

Requires: boto3, openpyxl, psycopg2
"""

import os
import sys
import argparse
from io import BytesIO

import boto3
import openpyxl
import psycopg2
from psycopg2.extras import RealDictCursor


def read_xlsx_base_costs(bucket, keys):
    """Read {sku: base_cost} from pricefeed xlsx files in S3."""
    s3 = boto3.client('s3', region_name='us-east-1')
    costs = {}

    for s3_key in keys:
        print(f"  Reading s3://{bucket}/{s3_key}")
        obj = s3.get_object(Bucket=bucket, Key=s3_key)
        wb = openpyxl.load_workbook(
            filename=BytesIO(obj['Body'].read()),
            read_only=True, data_only=True
        )

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = {str(c).strip().lower(): i for i, c in enumerate(header_row) if c}

            sku_idx = headers.get('sku')
            cost_idx = headers.get('base_cost')
            if sku_idx is None or cost_idx is None:
                continue

            count = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                sku = row[sku_idx]
                cost = row[cost_idx]
                if sku is None or cost is None:
                    continue
                try:
                    costs[str(sku).strip()] = float(cost)
                    count += 1
                except (ValueError, TypeError):
                    continue

            print(f"    Sheet '{sheet_name}': {count} base_cost values")

        wb.close()

    return costs


def read_rds_cost_gbp(connection, table_name='cardrush_link'):
    """Read {sku: cost_gbp} from RDS."""
    with connection.cursor() as cursor:
        cursor.execute(f"""
            SELECT sku, cost_gbp
            FROM {table_name}
            WHERE cost_gbp IS NOT NULL AND cost_gbp > 0
        """)
        return {row['sku']: float(row['cost_gbp']) for row in cursor.fetchall()}


def main():
    parser = argparse.ArgumentParser(description='Verify landed-cost model against pricefeed xlsx')
    parser.add_argument('--bucket', default='pricedata-tcg', help='S3 bucket')
    parser.add_argument('--keys', default='pricefeed/onepiece_pricefeed.xlsx,pricefeed/pokemon_pricefeed.xlsx',
                        help='Comma-separated S3 keys')
    parser.add_argument('--table', default='cardrush_link', help='RDS table name')
    parser.add_argument('--rate', type=float, default=5.0, help='Shipping rate percent (default: 5)')
    parser.add_argument('--flat', type=float, default=1.00, help='Shipping flat fee GBP (default: 1.00)')
    args = parser.parse_args()

    keys = [k.strip() for k in args.keys.split(',') if k.strip()]
    shipping_rate = args.rate / 100.0
    shipping_flat = args.flat

    print("=" * 60)
    print("Landed Cost Model Verification")
    print(f"Model: landed = cost_gbp * {1 + shipping_rate:.2f} + {shipping_flat:.2f}")
    print("=" * 60)

    # 1. Read xlsx base_cost
    print("\n1. Reading base_cost from S3 xlsx...")
    xlsx_costs = read_xlsx_base_costs(args.bucket, keys)
    print(f"   Total: {len(xlsx_costs)} SKUs with base_cost")

    # 2. Read RDS cost_gbp
    print("\n2. Reading cost_gbp from RDS...")
    connection = psycopg2.connect(
        host=os.environ['PROXY_ENDPOINT'],
        database=os.environ.get('DATABASE_NAME', 'op_cardrush_link'),
        user=os.environ['DB_USER'],
        password=os.environ['DB_PASSWORD'],
        port=int(os.environ.get('DB_PORT', 5432)),
        cursor_factory=RealDictCursor,
        connect_timeout=10
    )
    rds_costs = read_rds_cost_gbp(connection, args.table)
    connection.close()
    print(f"   Total: {len(rds_costs)} SKUs with cost_gbp")

    # 3. Match SKUs
    common_skus = set(xlsx_costs.keys()) & set(rds_costs.keys())
    print(f"\n3. Matched: {len(common_skus)} SKUs in both sources")

    if len(common_skus) < 5:
        print("   ERROR: Too few matched SKUs for comparison")
        sys.exit(1)

    # 4. Compare model vs xlsx
    print(f"\n4. Comparing model predictions vs xlsx base_cost:")
    errors = []
    for sku in sorted(common_skus):
        predicted = rds_costs[sku] * (1 + shipping_rate) + shipping_flat
        actual = xlsx_costs[sku]
        error = actual - predicted
        pct_error = (error / actual * 100) if actual != 0 else 0
        errors.append((sku, rds_costs[sku], actual, predicted, error, pct_error))

    abs_errors = [abs(e[4]) for e in errors]
    mae = sum(abs_errors) / len(abs_errors)
    max_error = max(abs_errors)

    print(f"   MAE  = £{mae:.4f}")
    print(f"   Max  = £{max_error:.4f}")
    print(f"   N    = {len(errors)}")

    # Show worst outliers
    errors.sort(key=lambda e: abs(e[4]), reverse=True)
    print(f"\n   Top 10 deviations:")
    print(f"   {'SKU':<20} {'cost_gbp':>10} {'xlsx':>10} {'model':>10} {'error':>10} {'%err':>8}")
    for sku, cg, actual, pred, err, pct in errors[:10]:
        print(f"   {sku:<20} £{cg:>8.2f} £{actual:>8.2f} £{pred:>8.2f} £{err:>8.2f} {pct:>7.1f}%")

    print(f"\n" + "=" * 60)
    print(f"Current env vars:")
    print(f"  SHIPPING_RATE={args.rate:.0f}")
    print(f"  SHIPPING_FLAT_GBP={shipping_flat:.2f}")
    print(f"=" * 60)


if __name__ == '__main__':
    main()
