"""
Arbitrage & Competition Dashboard — Local CLI

Scans eBay competitors and generates a self-contained HTML report.
Bypasses RDS entirely: S3 pricefeed → daily_prices fallback → Browse API → HTML.

Data sources (all from S3, no RDS needed):
  - pricefeed/{game}_pricefeed.xlsx  — authoritative cost_gbp, selling_price, FX rate
  - daily_prices.xlsx                — cost_yen with backward fallback through date columns

Usage:
    python3.11 -m pricing.browse.dashboard                     # full scan, open in browser
    python3.11 -m pricing.browse.dashboard --max-skus 20       # limit for quick test
    python3.11 -m pricing.browse.dashboard --output report.html # custom output path
    python3.11 -m pricing.browse.dashboard --no-open            # don't auto-open browser
"""

import argparse
import json
import os
import sys
import tempfile
import time
import webbrowser
from datetime import datetime
from io import BytesIO

import boto3
import openpyxl
from dotenv import load_dotenv


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

S3_BUCKET = 'pricedata-tcg'

# Pricefeed files (authoritative: has pre-computed cost_gbp + selling_price)
PRICEFEED_KEYS = {
    'onepiece': 'pricefeed/onepiece_pricefeed.xlsx',
    'pokemon': 'pricefeed/pokemon_pricefeed.xlsx',
}
# Sheet within each pricefeed to use for eBay business prices
PRICEFEED_SHEET = 'ebay_business'

# Daily prices xlsx (cost_yen source, with date column fallback)
DAILY_PRICES_KEY = 'daily_prices.xlsx'
DAILY_PRICES_SHEETS = ('onepiece', 'pokemon')

# Classification thresholds (same as lambda_function.py)
ACQUISITION_THRESHOLD = 0.50
UNDERPRICED_THRESHOLD = 0.70


# ---------------------------------------------------------------------------
# SKU parser (copied from lambda_function.py to avoid psycopg2 import chain)
# ---------------------------------------------------------------------------

def parse_sku(sku):
    """Parse a SKU into game, set_code, card_number, language."""
    if not sku:
        return None
    parts = sku.split('-')
    if len(parts) < 4:
        return None
    game = parts[0]
    set_code = parts[1]
    card_number = parts[2]
    lang_code = parts[3]
    if game not in ('OP', 'PKMN'):
        return None
    language = 'Japanese' if lang_code == 'JP' else lang_code
    return {
        'game': game,
        'set_code': set_code,
        'card_number': card_number,
        'language': language,
    }


def classify_price_ratio(ratio):
    """Classify competitor price ratio vs our acquisition cost."""
    if ratio < ACQUISITION_THRESHOLD:
        return 'ACQUISITION'
    elif ratio < UNDERPRICED_THRESHOLD:
        return 'UNDERPRICED'
    elif ratio <= 1.0:
        return 'COMPETITIVE'
    else:
        return 'ABOVE'


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def _safe_float(val):
    """Convert value to float, return None on failure."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def load_pricefeed():
    """Load pre-computed prices from S3 pricefeed xlsx files.

    Reads the ebay_business sheet from each pricefeed file.
    Returns dict keyed by SKU: {sku: {cost_gbp, selling_price, gbp_to_jpy}}
    """
    bucket = os.environ.get('SOURCE_BUCKET', S3_BUCKET)
    s3 = boto3.client('s3')
    sku_data = {}

    for game, key in PRICEFEED_KEYS.items():
        print(f"Downloading s3://{bucket}/{key}...")
        try:
            obj = s3.get_object(Bucket=bucket, Key=key)
        except Exception as e:
            print(f"  Failed to download {key}: {e}")
            continue

        wb = openpyxl.load_workbook(filename=BytesIO(obj['Body'].read()), data_only=True)

        if PRICEFEED_SHEET not in wb.sheetnames:
            print(f"  Sheet '{PRICEFEED_SHEET}' not found in {key}")
            continue

        sheet = wb[PRICEFEED_SHEET]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = {}
        for idx, cell in enumerate(header_row):
            if cell is not None:
                headers[str(cell).strip().lower()] = idx

        sku_idx = headers.get('sku')
        cost_idx = headers.get('base_cost')
        sell_idx = headers.get('selling_price')
        fx_idx = headers.get('gbp_to_jpy')

        if sku_idx is None or cost_idx is None or sell_idx is None:
            print(f"  Missing required columns in {key}. Found: {list(headers.keys())}")
            continue

        count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            sku = row[sku_idx] if sku_idx < len(row) else None
            if not sku:
                continue
            sku = str(sku).strip()

            cost_gbp = _safe_float(row[cost_idx] if cost_idx < len(row) else None)
            selling = _safe_float(row[sell_idx] if sell_idx < len(row) else None)
            fx = _safe_float(row[fx_idx] if fx_idx is not None and fx_idx < len(row) else None)

            if cost_gbp and cost_gbp > 0 and selling and selling > 0:
                sku_data[sku] = {
                    'cost_gbp': cost_gbp,
                    'selling_price': selling,
                    'gbp_to_jpy': fx or 0,
                }
                count += 1

        print(f"  {game}: {count} SKUs with valid prices")

    print(f"Pricefeed total: {len(sku_data)} SKUs")
    return sku_data


def load_daily_prices():
    """Load cost_yen + card_number from daily_prices.xlsx with backward date fallback.

    For each SKU, scans date columns from most recent to oldest to find the
    last valid (numeric) cost_yen. This handles "Not Available" entries from
    failed scraper runs.

    Returns dict keyed by SKU: {sku: {cost_yen, cost_yen_date, card_number}}
    """
    bucket = os.environ.get('SOURCE_BUCKET', S3_BUCKET)
    key = os.environ.get('SOURCE_KEY', DAILY_PRICES_KEY)

    print(f"Downloading s3://{bucket}/{key}...")
    s3 = boto3.client('s3')
    obj = s3.get_object(Bucket=bucket, Key=key)
    content = obj['Body'].read()

    workbook = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)
    sku_data = {}

    for sheet_name in DAILY_PRICES_SHEETS:
        if sheet_name not in workbook.sheetnames:
            print(f"  Sheet '{sheet_name}' not found, skipping")
            continue

        sheet = workbook[sheet_name]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = {}
        for idx, cell in enumerate(header_row):
            if cell is not None:
                headers[str(cell).strip().lower()] = idx

        sku_idx = headers.get('sku')
        cn_idx = headers.get('card_number')
        if sku_idx is None:
            print(f"  Sheet '{sheet_name}': no 'sku' column, skipping")
            continue

        # Collect date columns sorted newest-first for backward fallback
        date_cols = []
        for col_name, idx in headers.items():
            if len(col_name) == 10 and col_name[4] == '-' and col_name[7] == '-':
                try:
                    datetime.strptime(col_name, '%Y-%m-%d')
                    date_cols.append((col_name, idx))
                except ValueError:
                    pass

        date_cols.sort(key=lambda x: x[0], reverse=True)  # newest first

        if not date_cols:
            print(f"  Sheet '{sheet_name}': no date columns found")
            continue

        print(f"  Sheet '{sheet_name}': {len(date_cols)} date columns "
              f"({date_cols[-1][0]} → {date_cols[0][0]})")

        count = 0
        fallback_count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            sku = row[sku_idx] if sku_idx < len(row) else None
            if not sku:
                continue
            sku = str(sku).strip()
            if not sku:
                continue

            # Scan date columns newest→oldest for first valid numeric price
            cost_yen = None
            cost_date = None
            for date_name, col_idx in date_cols:
                val = row[col_idx] if col_idx < len(row) else None
                parsed = _safe_float(val)
                if parsed is not None and parsed > 0:
                    cost_yen = parsed
                    cost_date = date_name
                    break

            card_number = ''
            if cn_idx is not None and cn_idx < len(row) and row[cn_idx]:
                card_number = str(row[cn_idx]).strip()

            if cost_yen:
                is_latest = (cost_date == date_cols[0][0])
                if not is_latest:
                    fallback_count += 1
                sku_data[sku] = {
                    'cost_yen': cost_yen,
                    'cost_yen_date': cost_date,
                    'card_number': card_number,
                }
                count += 1

        print(f"  {sheet_name}: {count} SKUs with valid cost_yen "
              f"({fallback_count} used fallback date)")

    print(f"Daily prices total: {len(sku_data)} SKUs")
    return sku_data


def merge_data(pricefeed, daily_prices):
    """Merge pricefeed (authoritative prices) with daily_prices (cost_yen + card info).

    Returns list of dicts ready for scanning.
    """
    rows = []
    for sku, pf in pricefeed.items():
        dp = daily_prices.get(sku, {})
        rows.append({
            'sku': sku,
            'cost_gbp': pf['cost_gbp'],
            'selling_price': pf['selling_price'],
            'gbp_to_jpy': pf['gbp_to_jpy'],
            'cost_yen': dp.get('cost_yen', pf['cost_gbp'] * pf['gbp_to_jpy'] if pf['gbp_to_jpy'] else 0),
            'cost_yen_date': dp.get('cost_yen_date', ''),
            'card_number': dp.get('card_number', ''),
        })
    return rows


# ---------------------------------------------------------------------------
# Browse API scan
# ---------------------------------------------------------------------------

def scan_all(client, rows, max_skus=0):
    """Scan Browse API for each SKU. Returns list of result dicts."""
    if max_skus:
        rows = rows[:max_skus]

    results = []
    total = len(rows)

    for i, row in enumerate(rows):
        sku = row['sku']
        cost_gbp = row['cost_gbp']
        selling_price = row['selling_price']
        parsed = parse_sku(sku)

        base = {
            'sku': sku,
            'card_number': row['card_number'],
            'set_code': parsed['set_code'] if parsed else '',
            'cost_gbp': cost_gbp,
            'selling_price': selling_price,
            'cost_yen': row.get('cost_yen', 0),
            'cost_yen_date': row.get('cost_yen_date', ''),
        }

        if not parsed:
            base['error'] = 'unparseable SKU'
            results.append(base)
            continue

        try:
            competitors = client.search_competitors(
                game=parsed['game'],
                set_code=parsed['set_code'],
                card_number=parsed['card_number'],
                language=parsed['language'],
                our_price=selling_price,
            )
        except Exception as e:
            base['error'] = str(e)
            results.append(base)
            continue

        if not competitors:
            base.update({
                'cheapest_price': None,
                'cheapest_seller': '',
                'cheapest_url': '',
                'cheapest_title': '',
                'ratio': None,
                'classification': 'NO_RESULTS',
            })
        else:
            cheapest = competitors[0]
            ratio = cheapest['price'] / cost_gbp if cost_gbp > 0 else None
            classification = classify_price_ratio(ratio) if ratio else 'ABOVE'
            base.update({
                'cheapest_price': cheapest['price'],
                'cheapest_seller': cheapest.get('seller', ''),
                'cheapest_url': cheapest.get('url', ''),
                'cheapest_title': cheapest.get('title', ''),
                'ratio': ratio,
                'classification': classification,
            })

        results.append(base)

        # Progress
        if (i + 1) % 25 == 0 or i + 1 == total:
            print(f"  Scanned {i + 1}/{total}")

    return results


# ---------------------------------------------------------------------------
# HTML generation
# ---------------------------------------------------------------------------

def generate_html(results, metadata):
    """Build a self-contained HTML report with embedded CSS/JS."""

    # Count classifications
    counts = {}
    errors = 0
    no_results = 0
    for r in results:
        if 'error' in r:
            errors += 1
        elif r.get('classification') == 'NO_RESULTS':
            no_results += 1
        else:
            cls = r.get('classification', 'UNKNOWN')
            counts[cls] = counts.get(cls, 0) + 1

    # Prepare JSON data for JS
    table_data = []
    for r in results:
        if 'error' in r:
            continue
        table_data.append({
            'sku': r['sku'],
            'card_number': r['card_number'],
            'set_code': r.get('set_code', ''),
            'cost_yen': round(r.get('cost_yen', 0)),
            'cost_yen_date': r.get('cost_yen_date', ''),
            'cost_gbp': round(r['cost_gbp'], 2) if r['cost_gbp'] else 0,
            'selling_price': round(r['selling_price'], 2) if r['selling_price'] else 0,
            'cheapest_price': round(r['cheapest_price'], 2) if r.get('cheapest_price') else None,
            'cheapest_seller': r.get('cheapest_seller', ''),
            'cheapest_url': r.get('cheapest_url', ''),
            'cheapest_title': r.get('cheapest_title', ''),
            'ratio': round(r['ratio'], 4) if r.get('ratio') else None,
            'classification': r.get('classification', ''),
        })

    data_json = json.dumps(table_data)
    scan_time = metadata.get('scan_time', datetime.now().isoformat())
    fx_rate = metadata.get('fx_rate', 0)
    total_scanned = len(results)

    # Collect unique set codes for filter
    set_codes = sorted({r.get('set_code', '') for r in results if r.get('set_code')})
    set_codes_json = json.dumps(set_codes)

    stale_count = metadata.get('stale_count', 0)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Cambridge TCG — Arbitrage Dashboard</title>
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
         background: #0f1117; color: #e0e0e0; padding: 24px; }}

  .header {{ margin-bottom: 24px; }}
  .header h1 {{ font-size: 1.5rem; color: #fff; margin-bottom: 4px; }}
  .header .meta {{ font-size: 0.85rem; color: #888; }}
  .header .meta .stale-warn {{ color: #f97316; }}

  .summary {{ display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 24px; }}
  .card {{ flex: 1; min-width: 160px; padding: 16px; border-radius: 8px;
           border: 1px solid #2a2d35; }}
  .card .count {{ font-size: 2rem; font-weight: 700; }}
  .card .label {{ font-size: 0.8rem; text-transform: uppercase; letter-spacing: 0.05em;
                  margin-top: 4px; color: #aaa; }}
  .card-acq {{ border-color: #ef4444; background: rgba(239,68,68,0.08); }}
  .card-acq .count {{ color: #ef4444; }}
  .card-under {{ border-color: #f97316; background: rgba(249,115,22,0.08); }}
  .card-under .count {{ color: #f97316; }}
  .card-comp {{ border-color: #3b82f6; background: rgba(59,130,246,0.08); }}
  .card-comp .count {{ color: #3b82f6; }}
  .card-above {{ border-color: #22c55e; background: rgba(34,197,94,0.08); }}
  .card-above .count {{ color: #22c55e; }}
  .card-stats {{ border-color: #6b7280; background: rgba(107,114,128,0.06); }}
  .card-stats .count {{ color: #9ca3af; }}

  .filters {{ display: flex; gap: 12px; flex-wrap: wrap; margin-bottom: 16px;
              align-items: center; }}
  .filters select, .filters input {{
    background: #1a1d27; border: 1px solid #2a2d35; color: #e0e0e0;
    padding: 8px 12px; border-radius: 6px; font-size: 0.9rem;
  }}
  .filters input {{ width: 260px; }}
  .filters select:focus, .filters input:focus {{ outline: none; border-color: #3b82f6; }}

  table {{ width: 100%; border-collapse: collapse; font-size: 0.85rem; }}
  thead {{ position: sticky; top: 0; z-index: 2; }}
  th {{ background: #1a1d27; padding: 10px 12px; text-align: left; cursor: pointer;
       border-bottom: 2px solid #2a2d35; white-space: nowrap; user-select: none; }}
  th:hover {{ color: #3b82f6; }}
  th .sort-arrow {{ margin-left: 4px; font-size: 0.7rem; }}
  td {{ padding: 8px 12px; border-bottom: 1px solid #1e2028; }}
  tr:hover td {{ background: rgba(59,130,246,0.06); }}

  .row-ACQUISITION td {{ border-left: 3px solid #ef4444; }}
  .row-UNDERPRICED td {{ border-left: 3px solid #f97316; }}
  .row-COMPETITIVE td {{ border-left: 3px solid #3b82f6; }}
  .row-ABOVE td {{ border-left: 3px solid #22c55e; }}
  .row-NO_RESULTS td {{ border-left: 3px solid #4b5563; opacity: 0.6; }}

  .cls-badge {{ padding: 2px 8px; border-radius: 4px; font-size: 0.75rem;
                font-weight: 600; display: inline-block; }}
  .cls-ACQUISITION {{ background: rgba(239,68,68,0.2); color: #ef4444; }}
  .cls-UNDERPRICED {{ background: rgba(249,115,22,0.2); color: #f97316; }}
  .cls-COMPETITIVE {{ background: rgba(59,130,246,0.2); color: #3b82f6; }}
  .cls-ABOVE {{ background: rgba(34,197,94,0.2); color: #22c55e; }}
  .cls-NO_RESULTS {{ background: rgba(107,114,128,0.2); color: #6b7280; }}

  a {{ color: #60a5fa; text-decoration: none; }}
  a:hover {{ text-decoration: underline; }}
  .text-right {{ text-align: right; }}
  .empty {{ color: #4b5563; font-style: italic; }}
  .stale {{ color: #f97316; font-size: 0.75rem; }}
</style>
</head>
<body>

<div class="header">
  <h1>Cambridge TCG — Arbitrage & Competition Dashboard</h1>
  <div class="meta">
    Scanned: {scan_time} &nbsp;|&nbsp;
    FX Rate: 1 GBP = {fx_rate:.2f} JPY (from pricefeed)
    {f'&nbsp;|&nbsp; <span class="stale-warn">{stale_count} SKUs using fallback cost_yen</span>' if stale_count else ''}
  </div>
</div>

<div class="summary">
  <div class="card card-acq">
    <div class="count">{counts.get('ACQUISITION', 0)}</div>
    <div class="label">Acquisition — strong buy signals</div>
  </div>
  <div class="card card-under">
    <div class="count">{counts.get('UNDERPRICED', 0)}</div>
    <div class="label">Underpriced — below our cost</div>
  </div>
  <div class="card card-comp">
    <div class="count">{counts.get('COMPETITIVE', 0)}</div>
    <div class="label">Competitive — at or near cost</div>
  </div>
  <div class="card card-above">
    <div class="count">{counts.get('ABOVE', 0)}</div>
    <div class="label">Above — above our cost</div>
  </div>
  <div class="card card-stats">
    <div class="count">{total_scanned}</div>
    <div class="label">Total scanned ({errors} errors, {no_results} no results)</div>
  </div>
</div>

<div class="filters">
  <select id="filterClass">
    <option value="">All Classifications</option>
    <option value="ACQUISITION">Acquisition</option>
    <option value="UNDERPRICED">Underpriced</option>
    <option value="COMPETITIVE">Competitive</option>
    <option value="ABOVE">Above</option>
    <option value="NO_RESULTS">No Results</option>
  </select>
  <select id="filterSet">
    <option value="">All Sets</option>
  </select>
  <input type="text" id="filterText" placeholder="Search SKU, card name, seller...">
</div>

<table>
<thead>
  <tr>
    <th data-col="sku">SKU <span class="sort-arrow"></span></th>
    <th data-col="card_number">Card <span class="sort-arrow"></span></th>
    <th data-col="cost_yen" class="text-right">Cost &yen; <span class="sort-arrow"></span></th>
    <th data-col="cost_gbp" class="text-right">Cost &pound; <span class="sort-arrow"></span></th>
    <th data-col="selling_price" class="text-right">We Sell &pound; <span class="sort-arrow"></span></th>
    <th data-col="cheapest_price" class="text-right">Cheapest &pound; <span class="sort-arrow"></span></th>
    <th data-col="ratio" class="text-right">Ratio <span class="sort-arrow"></span></th>
    <th data-col="classification">Class <span class="sort-arrow"></span></th>
    <th data-col="cheapest_seller">Seller <span class="sort-arrow"></span></th>
    <th>Link</th>
  </tr>
</thead>
<tbody id="tableBody"></tbody>
</table>

<script>
const DATA = {data_json};
const SET_CODES = {set_codes_json};

// Populate set code filter
const setSelect = document.getElementById('filterSet');
SET_CODES.forEach(s => {{
  const opt = document.createElement('option');
  opt.value = s;
  opt.textContent = s;
  setSelect.appendChild(opt);
}});

// Sort state
let sortCol = 'ratio';
let sortAsc = true;

function compare(a, b) {{
  let va = a[sortCol], vb = b[sortCol];
  if (va == null && vb == null) return 0;
  if (va == null) return 1;
  if (vb == null) return -1;
  if (typeof va === 'string') {{
    va = va.toLowerCase(); vb = (vb || '').toLowerCase();
    return sortAsc ? va.localeCompare(vb) : vb.localeCompare(va);
  }}
  return sortAsc ? va - vb : vb - va;
}}

function render() {{
  const clsFilter = document.getElementById('filterClass').value;
  const setFilter = document.getElementById('filterSet').value;
  const textFilter = document.getElementById('filterText').value.toLowerCase();

  let filtered = DATA.filter(r => {{
    if (clsFilter && r.classification !== clsFilter) return false;
    if (setFilter && r.set_code !== setFilter) return false;
    if (textFilter) {{
      const haystack = (r.sku + ' ' + r.card_number + ' ' + r.cheapest_seller + ' ' + r.cheapest_title).toLowerCase();
      if (!haystack.includes(textFilter)) return false;
    }}
    return true;
  }});

  filtered.sort(compare);

  const tbody = document.getElementById('tableBody');
  tbody.innerHTML = '';

  filtered.forEach(r => {{
    const tr = document.createElement('tr');
    tr.className = 'row-' + r.classification;

    const price = r.cheapest_price != null ? '&pound;' + r.cheapest_price.toFixed(2) : '<span class="empty">—</span>';
    const ratio = r.ratio != null ? (r.ratio * 100).toFixed(1) + '%' : '<span class="empty">—</span>';
    const link = r.cheapest_url ? '<a href="' + r.cheapest_url + '" target="_blank" rel="noopener">View</a>' : '';
    const today = new Date().toISOString().slice(0, 10);
    const yenDate = r.cost_yen_date && r.cost_yen_date !== today ? ' <span class="stale">' + r.cost_yen_date + '</span>' : '';

    tr.innerHTML = `
      <td>${{r.sku}}</td>
      <td>${{r.card_number}}</td>
      <td class="text-right">&yen;${{r.cost_yen.toLocaleString()}}${{yenDate}}</td>
      <td class="text-right">&pound;${{r.cost_gbp.toFixed(2)}}</td>
      <td class="text-right">&pound;${{r.selling_price.toFixed(2)}}</td>
      <td class="text-right">${{price}}</td>
      <td class="text-right">${{ratio}}</td>
      <td><span class="cls-badge cls-${{r.classification}}">${{r.classification}}</span></td>
      <td>${{r.cheapest_seller}}</td>
      <td>${{link}}</td>
    `;
    tbody.appendChild(tr);
  }});
}}

// Sort on header click
document.querySelectorAll('th[data-col]').forEach(th => {{
  th.addEventListener('click', () => {{
    const col = th.dataset.col;
    if (sortCol === col) {{
      sortAsc = !sortAsc;
    }} else {{
      sortCol = col;
      sortAsc = true;
    }}
    // Update arrow indicators
    document.querySelectorAll('th .sort-arrow').forEach(s => s.textContent = '');
    th.querySelector('.sort-arrow').textContent = sortAsc ? '\\u25B2' : '\\u25BC';
    render();
  }});
}});

// Filter events
document.getElementById('filterClass').addEventListener('change', render);
document.getElementById('filterSet').addEventListener('change', render);
document.getElementById('filterText').addEventListener('input', render);

// Initial render — default sort by ratio ascending (cheapest competitors first)
document.querySelector('th[data-col="ratio"] .sort-arrow').textContent = '\\u25B2';
render();
</script>
</body>
</html>"""
    return html


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description='Arbitrage & Competition Dashboard — scan eBay competitors, generate HTML report'
    )
    parser.add_argument('--max-skus', type=int, default=0,
                        help='Limit number of SKUs to scan (0 = all)')
    parser.add_argument('--output', type=str, default=None,
                        help='Output HTML file path (default: temp file)')
    parser.add_argument('--no-open', action='store_true',
                        help='Do not auto-open browser')
    args = parser.parse_args()

    # Load .env for credentials
    load_dotenv()

    # Add browse/ dir to sys.path so BrowseClient resolves
    browse_dir = os.path.dirname(os.path.abspath(__file__))
    if browse_dir not in sys.path:
        sys.path.insert(0, browse_dir)

    from browse_client import BrowseClient

    # Step 1: Load pricefeed (authoritative cost_gbp + selling_price)
    pricefeed = load_pricefeed()
    if not pricefeed:
        print("No pricefeed data found in S3. Aborting.")
        sys.exit(1)

    # Step 2: Load daily_prices (cost_yen with backward fallback + card_number)
    daily_prices = load_daily_prices()

    # Step 3: Merge
    rows = merge_data(pricefeed, daily_prices)
    print(f"\nMerged: {len(rows)} SKUs ready for scanning")

    # Count stale cost_yen (using fallback date)
    today = datetime.now().strftime('%Y-%m-%d')
    stale_count = sum(1 for r in rows if r['cost_yen_date'] and r['cost_yen_date'] != today)
    if stale_count:
        print(f"  {stale_count} SKUs using fallback cost_yen (latest scrape unavailable)")

    # Get FX rate from pricefeed (all rows should have the same rate)
    fx_rate = next((r['gbp_to_jpy'] for r in rows if r.get('gbp_to_jpy', 0) > 0), 0)

    # Step 4: Scan Browse API
    print("\nInitializing eBay Browse API client...")
    client = BrowseClient()
    print(f"Scanning {args.max_skus or len(rows)} SKUs...\n")
    start = time.time()
    results = scan_all(client, rows, max_skus=args.max_skus)
    elapsed = time.time() - start
    print(f"\nScan complete in {elapsed:.1f}s")

    # Step 5: Generate HTML
    metadata = {
        'scan_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'fx_rate': fx_rate,
        'stale_count': stale_count,
    }
    html = generate_html(results, metadata)

    # Write output
    if args.output:
        output_path = args.output
    else:
        fd, output_path = tempfile.mkstemp(suffix='.html', prefix='dashboard_')
        os.close(fd)

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"Report written to {output_path}")

    # Open in browser
    if not args.no_open:
        webbrowser.open(f'file://{os.path.abspath(output_path)}')


if __name__ == '__main__':
    main()
