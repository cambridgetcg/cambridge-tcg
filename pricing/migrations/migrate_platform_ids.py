"""
Data migration: Extract eBay item numbers from pricefeed xlsx files
and populate the new platform identifier columns in RDS.

Reads ebay_business sheet from pricefeed xlsx files,
matches rows by SKU, and writes ebay_item_number_business
into the cardrush_link table.

Usage:
    python migrate_platform_ids.py

Environment Variables Required:
    - PROXY_ENDPOINT: RDS Proxy endpoint
    - DB_USER: Database username
    - DB_PASSWORD: Database password
    - DATABASE_NAME: Database name (default: op_cardrush_link)
    - S3_BUCKET: S3 bucket containing pricefeed xlsx files
    - PRICEFEED_KEYS: Comma-separated S3 keys (e.g., "pricefeed_pokemon.xlsx,pricefeed_onepiece.xlsx")

Can also be run as a Lambda handler.
"""

import os
import re
import json
import boto3
import psycopg2
from psycopg2.extras import RealDictCursor, execute_batch
import openpyxl
from io import BytesIO
from datetime import datetime


def _safe_table_name(name):
    """Validate table name to prevent SQL injection."""
    if not re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*$', name):
        raise ValueError(f"Invalid table name: {name}")
    return name


def get_db_connection():
    """Connect to database through RDS Proxy"""
    return psycopg2.connect(
        host=os.environ['PROXY_ENDPOINT'],
        database=os.environ.get('DATABASE_NAME', 'op_cardrush_link'),
        user=os.environ['DB_USER'],
        password=os.environ['DB_PASSWORD'],
        port=int(os.environ.get('DB_PORT', 5432)),
        cursor_factory=RealDictCursor,
        connect_timeout=10
    )


def extract_item_numbers(file_content, sheet_name, sku_col='sku', item_number_col='ebay_item_number'):
    """
    Extract SKU -> item_number mapping from an xlsx sheet.

    Returns dict: {sku: item_number}
    """
    mapping = {}

    try:
        workbook = openpyxl.load_workbook(filename=BytesIO(file_content), data_only=True)
    except Exception as e:
        print(f"  Error loading workbook: {e}")
        return mapping

    if sheet_name not in workbook.sheetnames:
        print(f"  Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}")
        return mapping

    sheet = workbook[sheet_name]

    # Build header mapping (case-insensitive)
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {str(cell).strip().lower(): idx for idx, cell in enumerate(header_row) if cell}

    sku_idx = headers.get(sku_col.lower())
    item_idx = headers.get(item_number_col.lower())

    if sku_idx is None or item_idx is None:
        print(f"  Missing columns. Need '{sku_col}' and '{item_number_col}'. Found: {list(headers.keys())}")
        return mapping

    for row in sheet.iter_rows(min_row=2, values_only=True):
        sku = row[sku_idx]
        item_number = row[item_idx]
        if sku is not None and item_number is not None:
            mapping[str(sku).strip()] = str(item_number).strip()

    return mapping


def lambda_handler(event, context):
    """
    Lambda handler for platform ID migration.

    Event parameters (optional):
    - bucket: Override S3 bucket
    - pricefeed_keys: List of S3 keys to process
    - dry_run: If true, log but don't update
    - ebay_business_sheet: Sheet name for eBay business data (default: ebay_business)
    - sku_column: Column name for SKU (default: sku)
    - item_number_column: Column name for item number (default: ebay_item_number)
    """
    print("=" * 60)
    print("Platform ID Migration: xlsx -> RDS")
    print(f"Started: {datetime.now()}")
    print("=" * 60)

    s3 = boto3.client('s3')
    connection = None

    try:
        bucket = event.get('bucket', os.environ.get('S3_BUCKET'))
        pricefeed_keys = event.get('pricefeed_keys',
                                    os.environ.get('PRICEFEED_KEYS', '').split(','))
        pricefeed_keys = [k.strip() for k in pricefeed_keys if k.strip()]

        dry_run = event.get('dry_run', False)
        ebay_business_sheet = event.get('ebay_business_sheet', 'ebay_business')
        sku_column = event.get('sku_column', 'sku')
        item_number_column = event.get('item_number_column', 'ebay_item_number')

        print(f"Bucket: {bucket}")
        print(f"Files: {pricefeed_keys}")
        print(f"Dry run: {dry_run}")

        if not bucket or not pricefeed_keys:
            return {
                'statusCode': 400,
                'body': json.dumps({
                    'success': False,
                    'error': 'Missing S3_BUCKET or PRICEFEED_KEYS'
                })
            }

        # Aggregate mappings from all pricefeed files
        business_mapping = {}  # sku -> ebay_item_number

        for s3_key in pricefeed_keys:
            print(f"\nProcessing: s3://{bucket}/{s3_key}")

            try:
                obj = s3.get_object(Bucket=bucket, Key=s3_key)
                file_content = obj['Body'].read()
            except Exception as e:
                print(f"  Error fetching {s3_key}: {e}")
                continue

            # Extract from ebay_business sheet
            biz = extract_item_numbers(file_content, ebay_business_sheet, sku_column, item_number_column)
            print(f"  {ebay_business_sheet}: {len(biz)} item numbers found")
            business_mapping.update(biz)

        print(f"\nTotal mappings - Business: {len(business_mapping)}")

        if not business_mapping:
            return {
                'statusCode': 200,
                'body': json.dumps({
                    'success': True,
                    'message': 'No item numbers found in pricefeed files'
                })
            }

        # Connect to RDS and update
        print(f"\nConnecting to database...")
        connection = get_db_connection()
        table_name = _safe_table_name(os.environ.get('TABLE_NAME', 'cardrush_link'))

        business_updated = 0
        unmatched_skus = []

        with connection.cursor() as cursor:
            # Update ebay_item_number_business
            if business_mapping:
                business_data = [(item_num, sku) for sku, item_num in business_mapping.items()]

                if not dry_run:
                    execute_batch(
                        cursor,
                        f"UPDATE {table_name} SET ebay_item_number_business = %s WHERE sku = %s",
                        business_data,
                        page_size=100
                    )
                    business_updated = cursor.rowcount
                else:
                    business_updated = len(business_data)

                print(f"eBay Business: {'would update' if dry_run else 'updated'} {business_updated} rows")

            if not dry_run:
                connection.commit()

            # Check for unmatched SKUs
            all_skus = list(business_mapping.keys())
            cursor.execute(
                f"SELECT sku FROM {table_name} WHERE sku = ANY(%s)",
                (all_skus,)
            )
            matched_skus = {row['sku'] for row in cursor.fetchall()}
            unmatched_skus = [sku for sku in all_skus if sku not in matched_skus]

            if unmatched_skus:
                print(f"\nWARNING: {len(unmatched_skus)} SKUs not found in RDS:")
                for sku in unmatched_skus[:20]:
                    print(f"  - {sku}")
                if len(unmatched_skus) > 20:
                    print(f"  ... and {len(unmatched_skus) - 20} more")

        result = {
            'success': True,
            'dry_run': dry_run,
            'business': {
                'items_in_xlsx': len(business_mapping),
                'rows_updated': business_updated
            },
            'unmatched_skus': unmatched_skus,
            'unmatched_count': len(unmatched_skus),
            'timestamp': datetime.now().isoformat()
        }

        print("\n" + "=" * 60)
        print("Migration complete" + (" (DRY RUN)" if dry_run else ""))
        print(f"Business: {business_updated} rows")
        print(f"Unmatched SKUs: {len(unmatched_skus)}")
        print("=" * 60)

        return {
            'statusCode': 200,
            'body': json.dumps(result, default=str)
        }

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

        if connection:
            connection.rollback()

        return {
            'statusCode': 500,
            'body': json.dumps({'success': False, 'error': str(e)})
        }

    finally:
        if connection:
            connection.close()
            print("Database connection closed")


if __name__ == "__main__":
    # Local testing
    os.environ['PROXY_ENDPOINT'] = 'your-proxy-endpoint'
    os.environ['DB_USER'] = 'your-username'
    os.environ['DB_PASSWORD'] = 'your-password'
    os.environ['S3_BUCKET'] = 'your-bucket'
    os.environ['PRICEFEED_KEYS'] = 'pricefeed_pokemon.xlsx,pricefeed_onepiece.xlsx'

    result = lambda_handler({'dry_run': True}, {})
    print(json.dumps(json.loads(result['body']), indent=2))
