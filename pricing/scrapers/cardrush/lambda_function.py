import os
import re
import requests
from bs4 import BeautifulSoup
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
import boto3
import psycopg2
from psycopg2.extras import execute_batch
from requests.adapters import HTTPAdapter, Retry
from monitoring.metrics import record_pipeline_run

# --- Table name validation ---

def _safe_table_name(name):
    """Validate table name to prevent SQL injection."""
    if not re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*$', name):
        raise ValueError(f"Invalid table name: {name}")
    return name

# --- Database ---

def get_db_connection():
    """Connect to database through RDS Proxy"""
    connection = psycopg2.connect(
        host=os.environ['PROXY_ENDPOINT'],
        database=os.environ.get('DATABASE_NAME', 'op_cardrush_link'),
        user=os.environ['DB_USER'],
        password=os.environ['DB_PASSWORD'],
        port=int(os.environ.get('DB_PORT', 5432)),
        connect_timeout=10
    )
    return connection

# --- Helper Functions ---

def get_sheet_links_and_skus(sheet, sku_column_name):
    """
    Returns a list of tuples (row, url, sku) from the given sheet.
    Reads both the 'cardrush' URL column and the SKU column per row.
    """
    cardrush_col = None
    sku_col = None
    results = []

    # Identify column indices from header row (row 1)
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if cell_value:
            header_lower = str(cell_value).strip().lower()
            if header_lower == 'cardrush':
                cardrush_col = col
            if header_lower == sku_column_name.lower():
                sku_col = col

    if not cardrush_col:
        print(f"Sheet '{sheet.title}': 'cardrush' column not found.")
        return results

    if not sku_col:
        print(f"Sheet '{sheet.title}': '{sku_column_name}' column not found. RDS write will be skipped for this sheet.")

    # Gather links + SKUs from row 2 onward
    for row in range(2, sheet.max_row + 1):
        url = sheet.cell(row=row, column=cardrush_col).value
        if url:
            sku = None
            if sku_col:
                sku_val = sheet.cell(row=row, column=sku_col).value
                if sku_val:
                    sku = str(sku_val).strip()
            results.append((row, url, sku))
    return results

def find_empty_column(sheet):
    """
    Finds the first entirely empty column in the sheet.
    """
    for col in range(1, sheet.max_column + 2):
        if all(sheet.cell(row=row, column=col).value is None for row in range(1, sheet.max_row + 1)):
            return col
    return sheet.max_column + 1

def get_price_and_stock(session, url):
    """
    Fetch the price and stock count from the given URL.
    Returns (price, stock) where price is int or "Not Available",
    and stock is int or None.
    """
    try:
        response = session.get(url)
        response.raise_for_status()
        html = response.text
        soup = BeautifulSoup(html, "html.parser")

        # --- Price ---
        price = "Not Available"
        price_element = soup.find("span", id='pricech')
        if price_element:
            price_text = price_element.text.strip()
            price_digits = re.sub(r'\D', '', price_text)
            if price_digits.isdigit():
                price = int(price_digits)

        # --- Stock ---
        stock = None
        stock_match = re.search(r'在庫数\s*(\d+)', html)
        if stock_match:
            stock = int(stock_match.group(1))
        elif '在庫なし' in html:
            stock = 0

        return (price, stock)
    except Exception as e:
        print(f"Error fetching price/stock for {url}: {e}")
        return ("Not Available", None)


def get_stock_only(session, url):
    """Fetch only stock count from a CardRush product page."""
    try:
        response = session.get(url)
        response.raise_for_status()
        html = response.text
        stock_match = re.search(r'在庫数\s*(\d+)', html)
        if stock_match:
            return int(stock_match.group(1))
        if '在庫なし' in html:
            return 0
        return None
    except Exception as e:
        print(f"Error fetching stock for {url}: {e}")
        return None

# --- Main Lambda Handler ---

def lambda_handler(event, context):
    s3 = boto3.client('s3')

    # Environment variables
    source_bucket = os.environ.get('SOURCE_BUCKET')
    source_key = os.environ.get('SOURCE_KEY')
    sku_column_name = os.environ.get('SKU_COLUMN_NAME', 'sku')
    table_name = _safe_table_name(os.environ.get('TABLE_NAME', 'cardrush_link'))

    # Local path (Lambda's /tmp)
    local_file_path = '/tmp/pkmn_price_data.xlsx'

    # Download the Excel file from S3
    try:
        response = s3.get_object(Bucket=source_bucket, Key=source_key)
        with open(local_file_path, 'wb') as f:
            f.write(response['Body'].read())
        print(f"Downloaded {source_key} from bucket {source_bucket} to {local_file_path}")
    except Exception as e:
        error_msg = f"Error downloading file: {e}"
        print(error_msg)
        return {'statusCode': 500, 'body': error_msg}

    # Set today's date for header
    date_str = datetime.today().strftime('%Y-%m-%d')

    # Load workbook once with openpyxl
    try:
        workbook = load_workbook(local_file_path)
        sheet_names = workbook.sheetnames
        print(f"Found sheets: {sheet_names}")
    except Exception as e:
        error_msg = f"Error loading workbook: {e}"
        print(error_msg)
        return {'statusCode': 500, 'body': error_msg}

    # Prepare tasks: For each sheet that has a 'cardrush' column, record (sheet, row, url, sku)
    tasks = []  # Each task is a dict with: sheet_title, row, url, sku
    sheet_targets = {}  # Map sheet title to its target (empty) column where prices will be written

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        # Skip the Pricing sheet here; we'll handle it later.
        if sheet_name.lower() == "pricing":
            continue
        rows = get_sheet_links_and_skus(sheet, sku_column_name)
        if rows:
            target_col = find_empty_column(sheet)
            sheet_targets[sheet_name] = target_col
            for row, url, sku in rows:
                tasks.append({
                    'sheet': sheet_name,
                    'row': row,
                    'url': url,
                    'sku': sku
                })
        else:
            print(f"Sheet '{sheet_name}' has no links to process.")
    
    # Use a single requests session for connection pooling and retry mechanism
    session = requests.Session()
    retries = Retry(total=3, backoff_factor=1, status_forcelist=[429, 500, 502, 503, 504])
    adapter = HTTPAdapter(max_retries=retries)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    
    # Concurrently fetch prices and stock for all tasks
    results = {}  # Key: (sheet, row) -> (price, stock)
    with ThreadPoolExecutor(max_workers=20) as executor:
        future_to_task = {executor.submit(get_price_and_stock, session, task['url']): task for task in tasks}
        for future in as_completed(future_to_task):
            task = future_to_task[future]
            try:
                price, stock = future.result()
            except Exception as e:
                print(f"Error in task for URL {task['url']}: {e}")
                price, stock = "Not Available", None
            results[(task['sheet'], task['row'])] = (price, stock)
    
    # Update each sheet in the workbook with the fetched prices
    for sheet_name in sheet_targets:
        sheet = workbook[sheet_name]
        target_col = sheet_targets[sheet_name]
        # Write header with date in the empty column (first row)
        sheet.cell(row=1, column=target_col, value=date_str)
        
        # Update prices for each row that had a link (stock not written to Excel)
        for row in range(2, sheet.max_row + 1):
            key = (sheet_name, row)
            if key in results:
                price, _stock = results[key]
                sheet.cell(row=row, column=target_col, value=price)
    
    # Optionally update the Pricing sheet: 
    # This example writes the prices from the first processed sheet to Pricing sheet's column B.
    if "Pricing" in workbook.sheetnames:
        pricing_sheet = workbook["Pricing"]
        if sheet_targets:
            first_sheet = list(sheet_targets.keys())[0]
            sheet = workbook[first_sheet]
            target_col = sheet_targets[first_sheet]
            pricing_sheet.cell(row=1, column=2, value=date_str)
            for row in range(2, sheet.max_row + 1):
                pricing_sheet.cell(row=row, column=2, value=sheet.cell(row=row, column=target_col).value)
        else:
            print("No sheet available to update the Pricing sheet.")
    
    # Save the updated workbook once
    try:
        workbook.save(local_file_path)
        print("Workbook updated and saved locally.")
    except Exception as e:
        error_msg = f"Error saving workbook: {e}"
        print(error_msg)
        return {'statusCode': 500, 'body': error_msg}
    
    # --- RDS Write: Update price_yen by SKU (before S3, so failure prevents inconsistency) ---
    rds_updated = 0
    rds_skipped_no_sku = 0
    rds_skipped_no_price = 0
    connection = None

    # Build list of (price_yen, cardrush_stock, sku) tuples from scraped results
    rds_updates = []
    for (sheet_name, row), (price, stock) in results.items():
        # Find the matching task to get the SKU
        task = next((t for t in tasks if t['sheet'] == sheet_name and t['row'] == row), None)
        if not task or not task.get('sku'):
            rds_skipped_no_sku += 1
            continue
        if not isinstance(price, int):
            rds_skipped_no_price += 1
            continue
        rds_updates.append((price, stock, task['sku']))

    if rds_updates:
        try:
            print(f"\nConnecting to RDS for price_yen + stock update...")
            connection = get_db_connection()

            with connection.cursor() as cursor:
                execute_batch(
                    cursor,
                    f"UPDATE {table_name} SET price_yen = %s, cardrush_stock = %s WHERE sku = %s",
                    rds_updates,
                    page_size=100
                )
                rds_updated = cursor.rowcount
                connection.commit()

            print(f"RDS: updated price_yen + cardrush_stock for {rds_updated} rows")
            print(f"RDS: skipped {rds_skipped_no_sku} rows (no SKU), {rds_skipped_no_price} rows (no price)")

            # Log SKUs that were sent but didn't match any RDS row
            if rds_updated < len(rds_updates):
                print(f"RDS: {len(rds_updates) - rds_updated} SKUs not found in database (new products?)")

        except Exception as e:
            print(f"RDS write error: {e}")
            if connection:
                connection.rollback()
                record_pipeline_run(connection, 'scraper', 'failure', 0, str(e))
                connection.close()
            return {'statusCode': 500, 'body': f"RDS write failed: {e}"}
    else:
        print("No valid (SKU, price) pairs for RDS update.")

    # Fail if we expected RDS updates but got none
    if rds_updates and rds_updated == 0:
        if connection:
            record_pipeline_run(connection, 'scraper', 'failure', 0,
                                f"{len(rds_updates)} SKUs sent but 0 rows updated")
            connection.close()
        return {
            'statusCode': 500,
            'body': f"FAILURE: {len(rds_updates)} SKUs sent to RDS but 0 rows updated. "
                    f"Check SKU column and table contents."
        }

    # --- A- pass: scrape price + stock from cardrush_url_subgrade pages ---
    subgrade_updated = 0
    if connection:
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    f"SELECT sku, cardrush_url_subgrade FROM {table_name} "
                    "WHERE cardrush_url_subgrade IS NOT NULL"
                )
                subgrade_rows = cursor.fetchall()

            if subgrade_rows:
                print(f"\nA- pass: fetching price+stock for {len(subgrade_rows)} subgrade URLs...")
                subgrade_results = {}  # sku -> (price, stock)
                with ThreadPoolExecutor(max_workers=20) as executor:
                    future_to_sku = {
                        executor.submit(get_price_and_stock, session, url): sku
                        for sku, url in subgrade_rows
                    }
                    for future in as_completed(future_to_sku):
                        sku = future_to_sku[future]
                        try:
                            price, stock = future.result()
                        except Exception as e:
                            print(f"Error fetching A- data for {sku}: {e}")
                            price, stock = "Not Available", None
                        subgrade_results[sku] = (price, stock)

                # Batch update price_yen_subgrade + cardrush_stock_subgrade
                sub_updates = []
                for sku, (price, stock) in subgrade_results.items():
                    p = price if isinstance(price, int) else None
                    if p is not None or stock is not None:
                        sub_updates.append((p, stock, sku))
                if sub_updates:
                    with connection.cursor() as cursor:
                        execute_batch(
                            cursor,
                            f"UPDATE {table_name} SET price_yen_subgrade = %s, cardrush_stock_subgrade = %s WHERE sku = %s",
                            sub_updates,
                            page_size=100
                        )
                        subgrade_updated = cursor.rowcount
                        connection.commit()
                    print(f"RDS: updated price_yen_subgrade + cardrush_stock_subgrade for {subgrade_updated} rows")
            else:
                print("No subgrade URLs found — skipping A- pass.")

        except Exception as e:
            print(f"WARNING: A- stock pass failed (non-fatal): {e}")
            try:
                connection.rollback()
            except Exception:
                pass

    # Record success in pipeline_runs
    if connection:
        record_pipeline_run(connection, 'scraper', 'success', rds_updated)
        connection.close()

    # Upload the updated Excel file back to S3 (after RDS success)
    try:
        with open(local_file_path, 'rb') as f:
            s3.put_object(Bucket=source_bucket, Key=source_key, Body=f.read())
        print(f"Uploaded updated file to s3://{source_bucket}/{source_key}")
    except Exception as e:
        # S3 is archive-only — RDS is the source of truth, so log warning and continue
        print(f"WARNING: S3 upload failed (RDS update succeeded): {e}")

    return {
        'statusCode': 200,
        'body': f"Price scraping completed on {date_str} for sheets: {sheet_names}. "
                f"RDS: {rds_updated} price_yen+stock rows written, "
                f"{subgrade_updated} A- stock rows written. S3 archive updated."
    }
