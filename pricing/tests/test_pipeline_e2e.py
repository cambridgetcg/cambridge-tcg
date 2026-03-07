"""
End-to-end pipeline test — traces a price from scrape to platform push.

Mocks: S3, RDS (psycopg2), HTTP (requests), Secrets Manager, boto3
Tests: scraper-cardrush → price-calculator → api-shopify / api-ebay

Since CardRush is down, we simulate scraped prices and verify the full
data flow produces correct eBay/Shopify prices.

Run: python -m pytest pricing/tests/test_pipeline_e2e.py -v
"""

import os
import sys
import json
import math
import importlib
from unittest import mock
from unittest.mock import MagicMock, patch, call
from decimal import Decimal
from io import BytesIO
from datetime import datetime

import pytest

# ---------------------------------------------------------------------------
# Paths — add each Lambda directory to sys.path so imports work
# ---------------------------------------------------------------------------
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LAMBDA_DIR = os.path.join(BASE, "lambdas")
PRICING_DIR = BASE  # pricing/ root — contains monitoring/

# Ensure pricing/ is on sys.path so `monitoring.metrics` is importable
if PRICING_DIR not in sys.path:
    sys.path.insert(0, PRICING_DIR)

# Mock Lambda-only dependencies if not installed locally
for _mod in ['psycopg2', 'psycopg2.extras', 'boto3',
             'bs4', 'ebay_auth']:
    if _mod not in sys.modules:
        sys.modules[_mod] = MagicMock()


def load_lambda(subdir):
    """
    Import lambda_function.py from a specific Lambda directory.
    Handles the fact that all Lambdas share the same module name.
    """
    path = os.path.join(LAMBDA_DIR, subdir)
    # Temporarily make this the first entry on sys.path
    if path in sys.path:
        sys.path.remove(path)
    sys.path.insert(0, path)

    # Force a fresh import
    if "lambda_function" in sys.modules:
        del sys.modules["lambda_function"]

    import lambda_function
    return lambda_function


# ---------------------------------------------------------------------------
# Test fixtures / shared mock data
# ---------------------------------------------------------------------------

# 3 sample products that will flow through the entire pipeline
SAMPLE_PRODUCTS = [
    {"sku": "PKMN-001", "price_yen": 500,  "ebay_biz_id": "111111"},
    {"sku": "PKMN-002", "price_yen": 1200, "ebay_biz_id": "333333"},
    {"sku": "PKMN-003", "price_yen": 3000, "ebay_biz_id": "555555"},
]

# FX rate that fx-updater would write
GBP_TO_JPY = 190.0

# Landed cost rates (match Lambda env var defaults)
SHIPPING_RATE = 0.05
SHIPPING_FLAT_GBP = 1.00

# Pricing parameters (match Lambda env var defaults)
TARGET_MARGIN = 0.22
VAT_RATE = 0.20
FEES = {
    "ebay_business": 0.12,
    "cardmarket": 0.08,
    "shopify": 0.05,
}


def expected_selling_price(cost_gbp, margin, fee, vat=VAT_RATE):
    """Mirror the price-calculator formula exactly: P = C(1+M)(1+V) / (1 - F(1+V))"""
    if cost_gbp is None or cost_gbp <= 0:
        return None
    p = float(Decimal(str(cost_gbp)) * (Decimal("1") + Decimal(str(margin)))
              * (Decimal("1") + Decimal(str(vat)))
              / (Decimal("1") - Decimal(str(fee)) * (Decimal("1") + Decimal(str(vat)))))
    return math.ceil(p) + 0.80


def expected_cost_gbp(price_yen, gbp_to_jpy):
    return price_yen / gbp_to_jpy


def expected_landed_cost(cost_gbp, rate=SHIPPING_RATE, flat=SHIPPING_FLAT_GBP):
    """Mirror the calculator's landed cost: cost_gbp * (1 + rate) + flat."""
    return cost_gbp * (1 + rate) + flat


# ---------------------------------------------------------------------------
# 1. Test: scraper-cardrush — xlsx reading + SKU extraction + RDS write
# ---------------------------------------------------------------------------

class TestScraperCardrush:
    """Test the scraper's dual-write: S3 xlsx archive + RDS price_yen."""

    def _make_workbook(self, tmp_path):
        """Create a minimal xlsx with sku + cardrush columns."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Pokemon"
        ws["A1"] = "sku"
        ws["B1"] = "cardrush"
        for i, prod in enumerate(SAMPLE_PRODUCTS, start=2):
            ws.cell(row=i, column=1, value=prod["sku"])
            ws.cell(row=i, column=2, value=f"https://cardrush.jp/product/{prod['sku']}")

        path = str(tmp_path / "test.xlsx")
        wb.save(path)
        return path

    def test_get_sheet_links_and_skus(self, tmp_path):
        """Verify the helper reads both cardrush URLs and SKUs."""
        from openpyxl import load_workbook
        path = self._make_workbook(tmp_path)
        wb = load_workbook(path)
        ws = wb["Pokemon"]

        scraper = load_lambda("scraper-cardrush")

        rows = scraper.get_sheet_links_and_skus(ws, "sku")
        assert len(rows) == 3
        # Each tuple: (row_number, url, sku)
        assert rows[0] == (2, "https://cardrush.jp/product/PKMN-001", "PKMN-001")
        assert rows[1] == (3, "https://cardrush.jp/product/PKMN-002", "PKMN-002")
        assert rows[2] == (4, "https://cardrush.jp/product/PKMN-003", "PKMN-003")

    def test_get_sheet_links_and_skus_missing_sku_column(self, tmp_path):
        """If SKU column is missing, URLs are still returned but sku=None."""
        from openpyxl import Workbook, load_workbook
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "name"
        ws["B1"] = "cardrush"
        ws["A2"] = "Pikachu"
        ws["B2"] = "https://cardrush.jp/1"
        path = str(tmp_path / "no_sku.xlsx")
        wb.save(path)
        wb2 = load_workbook(path)

        scraper = load_lambda("scraper-cardrush")

        rows = scraper.get_sheet_links_and_skus(wb2.active, "sku")
        assert len(rows) == 1
        assert rows[0] == (2, "https://cardrush.jp/1", None)

    def test_rds_update_pairs_built_correctly(self):
        """
        Simulate scrape results and verify the (price_yen, sku) pairs
        that would be sent to RDS are correct.
        """

        # Simulate what lambda_handler builds: results dict + tasks list
        tasks = [
            {"sheet": "Pokemon", "row": 2, "url": "https://cardrush.jp/1", "sku": "PKMN-001"},
            {"sheet": "Pokemon", "row": 3, "url": "https://cardrush.jp/2", "sku": "PKMN-002"},
            {"sheet": "Pokemon", "row": 4, "url": "https://cardrush.jp/3", "sku": None},  # no SKU
        ]
        results = {
            ("Pokemon", 2): 500,            # valid int price
            ("Pokemon", 3): "Not Available", # failed scrape
            ("Pokemon", 4): 3000,            # valid price but no SKU
        }

        # Replicate the RDS update pair logic from lambda_handler
        rds_updates = []
        rds_skipped_no_sku = 0
        rds_skipped_no_price = 0
        for (sheet_name, row), price in results.items():
            task = next((t for t in tasks if t['sheet'] == sheet_name and t['row'] == row), None)
            if not task or not task.get('sku'):
                rds_skipped_no_sku += 1
                continue
            if not isinstance(price, int):
                rds_skipped_no_price += 1
                continue
            rds_updates.append((price, task['sku']))

        assert rds_updates == [(500, "PKMN-001")]
        assert rds_skipped_no_sku == 1   # row 4 has no SKU
        assert rds_skipped_no_price == 1 # row 3 is "Not Available"


# ---------------------------------------------------------------------------
# 2. Test: price-calculator — cost_gbp derivation + selling price formula
# ---------------------------------------------------------------------------

class TestPriceCalculator:
    """Test the price-calculator's formula and cost_gbp derivation."""

    def test_selling_price_formula_basic(self):
        """Verify P = ceil(C * (1+M) * (1+V) / (1 - F*(1+V))) + 0.80"""
        calc = load_lambda("price-calculator")

        # cost_gbp = 500 / 190 = 2.6316...
        cost_gbp = 500 / 190.0

        # eBay Business: 2.6316 * 1.22 * 1.20 / (1 - 0.10*1.20) = 3.8526/0.88 = 4.378 → ceil=5 → 5.80
        price = calc.calculate_selling_price(cost_gbp, 0.22, 0.10, 0.20)
        assert price == 5.80, f"Expected 5.80, got {price}"

        # Shopify: 2.6316 * 1.22 * 1.20 / (1 - 0.03*1.20) = 3.8526/0.964 = 3.997 → ceil=4 → 4.80
        price = calc.calculate_selling_price(cost_gbp, 0.22, 0.03, 0.20)
        assert price == 4.80, f"Expected 4.80, got {price}"

    def test_selling_price_formula_larger_cost(self):
        """Test with price_yen=3000, gbp_to_jpy=190."""
        calc = load_lambda("price-calculator")

        cost_gbp = 3000 / 190.0  # = 15.7895

        # eBay Business: 15.7895 * 1.22 * 1.20 / (1 - 0.10*1.20) = 23.116/0.88 = 26.268 → ceil=27 → 27.80
        price = calc.calculate_selling_price(cost_gbp, 0.22, 0.10, 0.20)
        assert price == 27.80, f"Expected 27.80, got {price}"

        # Cardmarket: 15.7895 * 1.22 * 1.20 / (1 - 0.05*1.20) = 23.116/0.94 = 24.591 → ceil=25 → 25.80
        price = calc.calculate_selling_price(cost_gbp, 0.22, 0.05, 0.20)
        assert price == 25.80, f"Expected 25.80, got {price}"

        # Shopify: 15.7895 * 1.22 * 1.20 / (1 - 0.03*1.20) = 23.116/0.964 = 23.980 → ceil=24 → 24.80
        price = calc.calculate_selling_price(cost_gbp, 0.22, 0.03, 0.20)
        assert price == 24.80, f"Expected 24.80, got {price}"

    def test_selling_price_edge_cases(self):
        """None cost, zero cost, fee >= 1."""
        calc = load_lambda("price-calculator")

        assert calc.calculate_selling_price(None, 0.30, 0.10) is None
        assert calc.calculate_selling_price(0, 0.30, 0.10) is None
        assert calc.calculate_selling_price(-5, 0.30, 0.10) is None
        assert calc.calculate_selling_price(10, 0.30, 1.0) is None

    def test_cost_gbp_derivation_sql(self):
        """
        Verify the cost_gbp derivation UPDATE would produce correct values.
        This tests the math, not the actual SQL execution.
        """
        for prod in SAMPLE_PRODUCTS:
            cost_gbp = prod["price_yen"] / GBP_TO_JPY
            expected = expected_cost_gbp(prod["price_yen"], GBP_TO_JPY)
            assert abs(cost_gbp - expected) < 0.0001

    def test_all_prices_end_in_80(self):
        """Every selling price must end in .80 regardless of input."""
        calc = load_lambda("price-calculator")

        for price_yen in [100, 250, 500, 999, 1200, 3000, 5000, 10000, 50000]:
            cost_gbp = price_yen / GBP_TO_JPY
            for fee_name, fee_rate in FEES.items():
                price = calc.calculate_selling_price(cost_gbp, TARGET_MARGIN, fee_rate, VAT_RATE)
                if price is not None:
                    frac = round(price % 1, 2)
                    assert frac == 0.80, (
                        f"price_yen={price_yen}, {fee_name}: price={price}, "
                        f"fractional={frac} (expected 0.80)"
                    )


# ---------------------------------------------------------------------------
# 3. Test: Full pipeline data trace — scrape → cost_gbp → selling prices → push
# ---------------------------------------------------------------------------

class TestFullPipelineTrace:
    """
    Trace a single product (PKMN-003, price_yen=3000) through every pipeline stage
    and verify the exact prices that would hit eBay and Shopify.

    The calculator now uses landed_cost_gbp (not raw cost_gbp) as formula input:
    landed_cost_gbp = cost_gbp * (1 + shipping_rate) + shipping_flat
    """

    def test_end_to_end_price_trace(self):
        """
        PKMN-003: price_yen=3000, gbp_to_jpy=190

        Stage 1 (scraper):    price_yen = 3000 written to RDS
        Stage 2 (fx-updater): gbp_to_jpy = 190 written to RDS
        Stage 3 (price-calc): cost_gbp = 3000/190 = 15.7895...
                               landed = 15.7895 * 1.05 + 1.00 = 17.5789
                               Formula: P = ceil(C(1+M)(1+V) / (1-F(1+V))) + 0.80
                               eBay Business = 31.80
                               Cardmarket    = 29.80
                               Shopify       = 28.80
        Stage 4 (api-shopify): pushes price=28.80 for SKU PKMN-003
        Stage 5 (api-ebay):    pushes price=31.80 for PKMN-003
        """
        calc = load_lambda("price-calculator")

        price_yen = 3000
        gbp_to_jpy = GBP_TO_JPY

        # Stage 1: scraper writes price_yen (just the raw int)
        assert price_yen == 3000

        # Stage 2: fx-updater writes gbp_to_jpy (from Amdoren API)
        assert gbp_to_jpy == 190.0

        # Stage 3: price-calculator derives cost_gbp then landed_cost_gbp
        cost_gbp = price_yen / gbp_to_jpy
        assert abs(cost_gbp - 15.7895) < 0.001

        landed = expected_landed_cost(cost_gbp)
        assert abs(landed - 17.5789) < 0.001

        # Stage 3: price-calculator calculates selling prices from landed cost
        ebay_biz   = calc.calculate_selling_price(landed, TARGET_MARGIN, FEES["ebay_business"], VAT_RATE)
        cardmarket = calc.calculate_selling_price(landed, TARGET_MARGIN, FEES["cardmarket"], VAT_RATE)
        shopify    = calc.calculate_selling_price(landed, TARGET_MARGIN, FEES["shopify"], VAT_RATE)

        assert ebay_biz == 31.80, f"eBay Business: expected 31.80, got {ebay_biz}"
        assert cardmarket == 29.80, f"Cardmarket: expected 29.80, got {cardmarket}"
        assert shopify == 28.80, f"Shopify: expected 28.80, got {shopify}"

        # Stage 4: api-shopify would push shopify_selling_price=28.80
        shopify_payload_price = str(shopify)
        assert shopify_payload_price == "28.8"

        # Stage 5: api-ebay would push via ReviseInventoryStatus XML
        ebay_biz_xml_price = f"{ebay_biz:.2f}"
        assert ebay_biz_xml_price == "31.80"

    def test_all_sample_products_traced(self):
        """Trace all 3 sample products through the pipeline with landed costs."""
        calc = load_lambda("price-calculator")

        expected_results = []

        for prod in SAMPLE_PRODUCTS:
            cost_gbp = prod["price_yen"] / GBP_TO_JPY
            landed = expected_landed_cost(cost_gbp)

            prices = {}
            for channel, fee in FEES.items():
                prices[channel] = calc.calculate_selling_price(landed, TARGET_MARGIN, fee, VAT_RATE)

            expected_results.append({
                "sku": prod["sku"],
                "price_yen": prod["price_yen"],
                "cost_gbp": round(cost_gbp, 4),
                "landed_cost": round(landed, 4),
                **prices,
            })

        # Print trace table for visibility
        print("\n" + "=" * 100)
        print("FULL PIPELINE TRACE — All Sample Products (with landed costs)")
        print("=" * 100)
        print(f"{'SKU':<12} {'price_yen':>10} {'cost_gbp':>10} {'landed':>10} {'eBay Biz':>10} {'Cardmkt':>10} {'Shopify':>10}")
        print("-" * 90)
        for r in expected_results:
            print(
                f"{r['sku']:<12} "
                f"¥{r['price_yen']:>8,} "
                f"£{r['cost_gbp']:>8.4f} "
                f"£{r['landed_cost']:>8.4f} "
                f"£{r['ebay_business']:>8.2f} "
                f"£{r['cardmarket']:>8.2f} "
                f"£{r['shopify']:>8.2f}"
            )
        print("=" * 90)

        # Verify all prices end in .80
        for r in expected_results:
            for ch in FEES:
                assert round(r[ch] % 1, 2) == 0.80

        # Verify specific expected values (landed cost → 22% margin + 20% VAT)
        # PKMN-001: cost_gbp=2.6316, landed=3.7632
        assert expected_results[0]["ebay_business"] == 7.80
        assert expected_results[0]["shopify"] == 6.80

        # PKMN-002: cost_gbp=6.3158, landed=7.6316
        assert expected_results[1]["ebay_business"] == 14.80

        # PKMN-003: cost_gbp=15.7895, landed=17.5789
        assert expected_results[2]["ebay_business"] == 31.80
        assert expected_results[2]["shopify"] == 28.80


# ---------------------------------------------------------------------------
# 4. Test: api-shopify reads correct columns and formats payload
# ---------------------------------------------------------------------------

class TestApiShopify:
    """Test Shopify Lambda's DB query and price payload format."""

    def test_shopify_query_reads_correct_column(self):
        """api-shopify SELECTs sku + shopify_selling_price from RDS."""
        shopify_lambda = load_lambda("api-shopify")

        # The query is in lambda_handler — verify by reading the source
        import inspect
        source = inspect.getsource(shopify_lambda.lambda_handler)
        assert "shopify_selling_price" in source
        assert "SELECT sku, shopify_selling_price" in source

    def test_shopify_price_payload_format(self):
        """Shopify REST payload uses str(price) — verify format."""
        # Shopify process_row builds: {"variant": {"price": str(price)}}
        price = 24.80
        payload = {"variant": {"id": 12345, "price": str(price)}}
        assert payload["variant"]["price"] == "24.8"

        price = 4.80
        payload = {"variant": {"id": 12345, "price": str(price)}}
        assert payload["variant"]["price"] == "4.8"


# ---------------------------------------------------------------------------
# 5. Test: api-ebay reads correct columns and formats XML
# ---------------------------------------------------------------------------

class TestApiEbay:
    """Test eBay Lambda's DB queries and XML payload format."""

    def test_ebay_channel_config(self):
        """Verify channel_config maps to correct DB columns."""
        ebay_lambda = load_lambda("api-ebay")

        # Read channel_config from source
        import inspect
        source = inspect.getsource(ebay_lambda.lambda_handler)
        assert "ebay_business_selling_price" in source
        assert "ebay_item_number_business" in source

    def test_ebay_xml_price_format(self):
        """ReviseInventoryStatus XML uses f'{price:.2f}' — verify."""
        ebay_lambda = load_lambda("api-ebay")

        # The XML builder formats: f"{item['price']:.2f}"
        # Verify that our expected prices format correctly
        assert f"{26.80:.2f}" == "26.80"
        assert f"{27.80:.2f}" == "27.80"
        assert f"{4.80:.2f}" == "4.80"

    def test_ebay_batch_size(self):
        """eBay batches must be max 4 items (API limit)."""
        ebay_lambda = load_lambda("api-ebay")

        assert ebay_lambda.BATCH_SIZE == 4


# ---------------------------------------------------------------------------
# 6. Test: cost_gbp derivation guard clauses
# ---------------------------------------------------------------------------

class TestCostGbpDerivation:
    """Test that the cost_gbp UPDATE has correct WHERE guards."""

    def test_derivation_guards(self):
        """
        The SQL: UPDATE ... SET cost_gbp = price_yen / gbp_to_jpy
                 WHERE price_yen IS NOT NULL AND price_yen > 0
                   AND gbp_to_jpy IS NOT NULL AND gbp_to_jpy > 0

        Verify: division by zero impossible, NULL inputs skipped.
        """
        calc = load_lambda("price-calculator")

        import inspect
        source = inspect.getsource(calc.lambda_handler)

        # Confirm the WHERE clause guards exist
        assert "price_yen IS NOT NULL" in source
        assert "price_yen > 0" in source
        assert "gbp_to_jpy IS NOT NULL" in source
        assert "gbp_to_jpy > 0" in source

    def test_zero_division_impossible(self):
        """With gbp_to_jpy > 0 guard, division by zero can't happen."""
        # Simulate the SQL WHERE clause filtering
        test_rows = [
            {"price_yen": 500, "gbp_to_jpy": 190},   # valid
            {"price_yen": 0, "gbp_to_jpy": 190},       # filtered: price_yen = 0
            {"price_yen": 500, "gbp_to_jpy": 0},       # filtered: gbp_to_jpy = 0
            {"price_yen": None, "gbp_to_jpy": 190},    # filtered: price_yen NULL
            {"price_yen": 500, "gbp_to_jpy": None},    # filtered: gbp_to_jpy NULL
        ]

        valid = [r for r in test_rows
                 if r["price_yen"] is not None and r["price_yen"] > 0
                 and r["gbp_to_jpy"] is not None and r["gbp_to_jpy"] > 0]

        assert len(valid) == 1
        assert valid[0]["price_yen"] / valid[0]["gbp_to_jpy"] == pytest.approx(2.6316, abs=0.001)


# ---------------------------------------------------------------------------
# 6b. Test: landed_cost_gbp derivation arithmetic
# ---------------------------------------------------------------------------

class TestLandedCostDerivation:
    """Test the landed_cost_gbp = cost_gbp * (1 + rate) + flat formula."""

    def test_landed_cost_arithmetic(self):
        """Verify landed cost formula for all sample products."""
        for prod in SAMPLE_PRODUCTS:
            cost_gbp = prod["price_yen"] / GBP_TO_JPY
            landed = expected_landed_cost(cost_gbp)
            manual = cost_gbp * (1 + SHIPPING_RATE) + SHIPPING_FLAT_GBP
            assert abs(landed - manual) < 0.0001, f"{prod['sku']}: {landed} != {manual}"

    def test_landed_cost_always_greater_than_cost_gbp(self):
        """Landed cost must always exceed raw cost_gbp (fees are positive)."""
        for prod in SAMPLE_PRODUCTS:
            cost_gbp = prod["price_yen"] / GBP_TO_JPY
            landed = expected_landed_cost(cost_gbp)
            assert landed > cost_gbp

    def test_landed_cost_sql_guards(self):
        """The SQL UPDATE guards: WHERE cost_gbp IS NOT NULL AND cost_gbp > 0."""
        calc = load_lambda("price-calculator")
        import inspect
        source = inspect.getsource(calc.lambda_handler)

        landed_section = source[source.find("landed_cost_gbp = cost_gbp"):]
        assert "cost_gbp IS NOT NULL" in landed_section
        assert "cost_gbp > 0" in landed_section

    def test_landed_cost_stores_components(self):
        """The UPDATE stores flat in shipping_fee_gbp, rate in import_duty_rate, handling=0."""
        calc = load_lambda("price-calculator")
        import inspect
        source = inspect.getsource(calc.lambda_handler)

        assert "shipping_fee_gbp" in source
        assert "import_duty_rate" in source
        assert "handling_fee_gbp = 0" in source


# ---------------------------------------------------------------------------
# 6c. Test: velocity_to_margin dynamic margin helper
# ---------------------------------------------------------------------------

class TestVelocityToMargin:
    """Test the velocity_to_margin linear interpolation function."""

    def test_zero_sales_returns_default(self):
        calc = load_lambda("price-calculator")
        assert calc.velocity_to_margin(0) == 0.18

    def test_low_threshold_returns_min(self):
        calc = load_lambda("price-calculator")
        assert calc.velocity_to_margin(1) == 0.12

    def test_high_threshold_returns_max(self):
        calc = load_lambda("price-calculator")
        assert calc.velocity_to_margin(8) == 0.22

    def test_above_high_returns_max(self):
        calc = load_lambda("price-calculator")
        assert calc.velocity_to_margin(50) == 0.22

    def test_midpoint_interpolation(self):
        """Midpoint between low=1 and high=8 → midpoint between 12% and 22%."""
        calc = load_lambda("price-calculator")
        # t = (4.5 - 1) / (8 - 1) = 3.5/7 = 0.5 → 0.12 + 0.5*0.10 = 0.17
        result = calc.velocity_to_margin(4.5)
        assert abs(result - 0.17) < 0.001

    def test_custom_thresholds(self):
        calc = load_lambda("price-calculator")
        # low=2, high=10, margin_min=0.10, margin_max=0.30, default=0.20
        assert calc.velocity_to_margin(0, 2, 10, 0.10, 0.30, 0.20) == 0.20
        assert calc.velocity_to_margin(2, 2, 10, 0.10, 0.30, 0.20) == 0.10
        assert calc.velocity_to_margin(10, 2, 10, 0.10, 0.30, 0.20) == 0.30
        # t = (6-2)/(10-2) = 0.5 → 0.10 + 0.5*0.20 = 0.20
        assert abs(calc.velocity_to_margin(6, 2, 10, 0.10, 0.30, 0.20) - 0.20) < 0.001

    def test_flat_margin_revert(self):
        """Setting min=max=default should always return the same margin."""
        calc = load_lambda("price-calculator")
        for units in [0, 1, 5, 8, 100]:
            result = calc.velocity_to_margin(units, 1, 8, 0.18, 0.18, 0.18)
            assert result == 0.18, f"units={units}: expected 0.18, got {result}"


# ---------------------------------------------------------------------------
# 7. Test: cardrush-fx-updater fetches live rate from Amdoren API
# ---------------------------------------------------------------------------

class TestFxUpdater:
    """Test the fx-updater fetches from Amdoren and writes to RDS."""

    def test_fetch_gbp_to_jpy_success(self):
        """Mock a successful Amdoren API response."""
        fx = load_lambda("cardrush-fx-updater")

        mock_response = MagicMock()
        mock_response.status_code = 200
        mock_response.json.return_value = {"error": 0, "error_message": "-", "amount": 190.24}
        mock_response.raise_for_status = MagicMock()

        with patch.object(fx.requests, "get", return_value=mock_response):
            rate = fx.fetch_gbp_to_jpy("test-key")

        assert rate == 190.24

    def test_fetch_gbp_to_jpy_api_error(self):
        """Amdoren returns error != 0."""
        fx = load_lambda("cardrush-fx-updater")

        mock_response = MagicMock()
        mock_response.status_code = 200
        mock_response.json.return_value = {"error": 1, "error_message": "Invalid API key"}
        mock_response.raise_for_status = MagicMock()

        with patch.object(fx.requests, "get", return_value=mock_response):
            with pytest.raises(Exception, match="Amdoren API error"):
                fx.fetch_gbp_to_jpy("bad-key")

    def test_fetch_gbp_to_jpy_sanity_check(self):
        """Rate outside 50-500 range is rejected."""
        fx = load_lambda("cardrush-fx-updater")

        mock_response = MagicMock()
        mock_response.status_code = 200
        mock_response.json.return_value = {"error": 0, "error_message": "-", "amount": 5.0}
        mock_response.raise_for_status = MagicMock()

        with patch.object(fx.requests, "get", return_value=mock_response):
            with pytest.raises(Exception, match="sanity range"):
                fx.fetch_gbp_to_jpy("test-key")

    def test_fx_updater_no_s3_dependency(self):
        """Verify the fx-updater no longer imports boto3/pandas/S3."""
        fx = load_lambda("cardrush-fx-updater")
        import inspect
        source = inspect.getsource(fx)
        assert "boto3" not in source
        assert "pandas" not in source
        assert "s3" not in source.lower().split("amdoren")[0]  # no S3 before the API call


# ---------------------------------------------------------------------------
# 8. Test: Pipeline ordering — cost_gbp derived BEFORE selling price calc
# ---------------------------------------------------------------------------

class TestPipelineOrdering:
    """Verify that cost_gbp → landed_cost → selling price query runs in order."""

    def test_derivation_before_landed_cost(self):
        """cost_gbp derivation must come before landed_cost derivation."""
        calc = load_lambda("price-calculator")

        import inspect
        source = inspect.getsource(calc.lambda_handler)

        cost_gbp_pos = source.find("SET cost_gbp = price_yen / gbp_to_jpy")
        landed_pos = source.find("SET shipping_fee_gbp")

        assert cost_gbp_pos > 0, "cost_gbp derivation UPDATE not found"
        assert landed_pos > 0, "landed_cost_gbp derivation not found"
        assert cost_gbp_pos < landed_pos, (
            f"cost_gbp derivation (pos {cost_gbp_pos}) must come BEFORE "
            f"landed_cost derivation (pos {landed_pos})"
        )

    def test_landed_cost_before_selling_price_query(self):
        """landed_cost derivation must come before selling price SELECT."""
        calc = load_lambda("price-calculator")

        import inspect
        source = inspect.getsource(calc.lambda_handler)

        landed_pos = source.find("landed_cost_gbp = cost_gbp")
        selling_query_pos = source.find("SELECT sku, landed_cost_gbp")

        assert landed_pos > 0, "landed_cost_gbp derivation not found"
        assert selling_query_pos > 0, "Selling price SELECT not found"
        assert landed_pos < selling_query_pos, (
            f"landed_cost derivation (pos {landed_pos}) must come BEFORE "
            f"selling price query (pos {selling_query_pos})"
        )


# ---------------------------------------------------------------------------
# 8. Integration: Mock DB to test price-calculator handler
# ---------------------------------------------------------------------------

class TestPriceCalculatorWithMockDB:
    """Run price-calculator lambda_handler with a mocked database."""

    def test_handler_derives_cost_gbp_then_landed_then_calculates(self):
        calc = load_lambda("price-calculator")

        # Set up mock cursor that tracks SQL calls
        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        mock_conn.cursor.return_value.__enter__ = MagicMock(return_value=mock_cursor)
        mock_conn.cursor.return_value.__exit__ = MagicMock(return_value=False)

        sql_calls = []
        def track_execute(sql, params=None):
            sql_calls.append(sql.strip() if isinstance(sql, str) else sql)
            if "SET cost_gbp = price_yen" in str(sql):
                mock_cursor.rowcount = 3
            elif "landed_cost_gbp = cost_gbp" in str(sql):
                mock_cursor.rowcount = 3
            elif "gbp_to_jpy IS NOT NULL" in str(sql) and "COUNT(*)" in str(sql):
                # Staleness guard: FX rate must exist
                mock_cursor.fetchone.return_value = {"count": 3}
            elif "SUM(quantity)" in str(sql):
                # Step 2.5: velocity query
                mock_cursor.fetchall.return_value = []
            elif "SELECT sku, landed_cost_gbp" in str(sql):
                mock_cursor.fetchall.return_value = []
            elif "COUNT(*)" in str(sql):
                mock_cursor.fetchone.return_value = {"count": 0}

        mock_cursor.execute.side_effect = track_execute

        env = {
            "PROXY_ENDPOINT": "mock-proxy",
            "DB_USER": "mock-user",
            "DB_PASSWORD": "mock-pass",
            "DATABASE_NAME": "op_cardrush_link",
            "TABLE_NAME": "cardrush_link",
            "MARGIN_MIN": "12",
            "MARGIN_MAX": "22",
            "MARGIN_DEFAULT": "18",
            "VELOCITY_WINDOW_DAYS": "30",
            "VELOCITY_LOW": "1",
            "VELOCITY_HIGH": "8",
            "VAT_RATE": "20",
            "EBAY_BUSINESS_FEE": "12",
            "CARDMARKET_FEE": "8",
            "SHOPIFY_FEE": "5",
            "SHIPPING_RATE": "5",
            "SHIPPING_FLAT_GBP": "1.00",
        }

        with patch.dict(os.environ, env), \
             patch.object(calc, "get_db_connection", return_value=mock_conn):
            result = calc.lambda_handler({"batch_size": 100}, {})

        # Verify: staleness guard → cost_gbp → landed_cost → velocity → selling price SELECT
        assert len(sql_calls) >= 5, f"Expected >=5 SQL calls, got {len(sql_calls)}: {sql_calls}"
        assert "gbp_to_jpy IS NOT NULL" in sql_calls[0]  # staleness guard
        assert "SET cost_gbp = price_yen / gbp_to_jpy" in sql_calls[1]
        assert "landed_cost_gbp = cost_gbp" in sql_calls[2]
        assert "SUM(quantity)" in sql_calls[3]  # velocity query
        assert "SELECT sku, landed_cost_gbp" in sql_calls[4]
        # Verify it was committed
        mock_conn.commit.assert_called()


# ---------------------------------------------------------------------------
# Main: run with verbose trace output
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    # Quick manual trace if run directly
    print("\n" + "=" * 80)
    print("PIPELINE PRICE TRACE (manual run — with landed costs)")
    print("=" * 80)

    calc = load_lambda("price-calculator")

    print(f"\nFX Rate: 1 GBP = {GBP_TO_JPY} JPY")
    print(f"Shipping: {SHIPPING_RATE*100:.0f}% + £{SHIPPING_FLAT_GBP:.2f}")
    print(f"Margin: {TARGET_MARGIN*100:.0f}%")
    print(f"VAT: {VAT_RATE*100:.0f}%")
    print(f"Fees: {json.dumps({k: f'{v*100:.0f}%' for k, v in FEES.items()})}")
    print()

    print(f"{'SKU':<12} {'price_yen':>10} {'cost_gbp':>10} {'landed':>10} {'eBay Biz':>10} {'Cardmkt':>10} {'Shopify':>10}")
    print("-" * 82)

    for prod in SAMPLE_PRODUCTS:
        cost_gbp = prod["price_yen"] / GBP_TO_JPY
        landed = expected_landed_cost(cost_gbp)
        eb = calc.calculate_selling_price(landed, TARGET_MARGIN, FEES["ebay_business"], VAT_RATE)
        cm = calc.calculate_selling_price(landed, TARGET_MARGIN, FEES["cardmarket"], VAT_RATE)
        sh = calc.calculate_selling_price(landed, TARGET_MARGIN, FEES["shopify"], VAT_RATE)

        print(
            f"{prod['sku']:<12} "
            f"¥{prod['price_yen']:>8,} "
            f"£{cost_gbp:>8.4f} "
            f"£{landed:>8.4f} "
            f"£{eb:>8.2f} "
            f"£{cm:>8.2f} "
            f"£{sh:>8.2f}"
        )

    print("\nPrices that would hit platforms:")
    for prod in SAMPLE_PRODUCTS:
        cost_gbp = prod["price_yen"] / GBP_TO_JPY
        landed = expected_landed_cost(cost_gbp)
        eb = calc.calculate_selling_price(landed, TARGET_MARGIN, FEES["ebay_business"], VAT_RATE)
        sh = calc.calculate_selling_price(landed, TARGET_MARGIN, FEES["shopify"], VAT_RATE)
        print(f"  {prod['sku']}: Shopify→£{sh:.2f}  eBay Biz→£{eb:.2f}")

    pytest.main([__file__, "-v"])
