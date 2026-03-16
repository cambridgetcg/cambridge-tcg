"""CLI entry point for stock purchase list parsing and stock management.

Usage:
    # Seed stock from Zoho S3 export (run once to initialize)
    python -m stock.count.main import

    # Parse from stdin (e.g. pbpaste) with dry run
    pbpaste | python -m stock.count.main parse -s cardrush --dry-run

    # Parse from file and apply
    python -m stock.count.main parse -f purchase.txt -s cardrush

    # Check stock status
    python -m stock.count.main status
    python -m stock.count.main status --sku OP-OP05-001-JP

    # Export RDS stock to JSON backup
    python -m stock.count.main export -o stock_backup.json
"""

import argparse
import sys
from collections import defaultdict

from stock.count.models import SaleReduction, StockUpdate
from stock.count.parsers import PARSERS
from stock.count.resolver import SKUResolver
from stock.count.stock_store import StockStore


def cmd_parse(args):
    """Parse a supplier purchase list, resolve SKUs, and update stock."""
    # Read input
    if args.file:
        with open(args.file, 'r') as f:
            text = f.read()
    else:
        if sys.stdin.isatty():
            print("Error: No input. Pipe text via stdin or use -f FILE.", file=sys.stderr)
            sys.exit(1)
        text = sys.stdin.read()

    if not text.strip():
        print("Error: Empty input.", file=sys.stderr)
        sys.exit(1)

    # Parse
    parser_cls = PARSERS.get(args.supplier)
    if not parser_cls:
        print(f"Error: Unknown supplier '{args.supplier}'. Available: {list(PARSERS.keys())}", file=sys.stderr)
        sys.exit(1)

    parser = parser_cls()
    items = parser.parse(text)

    dropped = getattr(parser, 'dropped_lines', [])
    print(f"\nParsed {len(items)} line items from {args.supplier}")
    if dropped:
        print(f"  {len(dropped)} line(s) not recognized:")
        for line in dropped[:10]:
            print(f"    > {line[:80]}")
        if len(dropped) > 10:
            print(f"    ... and {len(dropped) - 10} more")

    if not items:
        print("No items parsed. Check input format.")
        return

    # Resolve SKUs
    resolver = SKUResolver()
    resolved = resolver.resolve(items)

    # Print resolution results
    print(f"\n{'='*70}")
    print(f"{'Card Number':<20} {'SKU':<30} {'Qty':>5} {'¥/unit':>8} {'Subtotal':>10}")
    print(f"{'='*70}")

    unresolved = []
    for r in resolved:
        p = r.parsed
        sku_display = r.sku or '???'
        marker = ''
        if r.ambiguous:
            marker = ' *'
        elif not r.resolved:
            marker = ' !'
        print(f"{p.card_number:<20} {sku_display:<30} {p.quantity:>5} {p.price_yen:>8,} {p.subtotal_yen:>10,}{marker}")

        for w in r.warnings:
            print(f"  WARNING: {w}")

        if not r.resolved:
            unresolved.append(r)

    # Aggregate by SKU
    sku_agg = defaultdict(lambda: {'qty': 0, 'cost': 0})
    for r in resolved:
        if r.sku:
            sku_agg[r.sku]['qty'] += r.parsed.quantity
            sku_agg[r.sku]['cost'] += r.parsed.subtotal_yen

    updates = [
        StockUpdate(sku=sku, quantity_to_add=v['qty'], cost_yen_total=v['cost'])
        for sku, v in sorted(sku_agg.items())
    ]

    total_qty = sum(u.quantity_to_add for u in updates)
    total_cost = sum(u.cost_yen_total for u in updates)

    print(f"\n{'─'*70}")
    print(f"{'Aggregated by SKU':^70}")
    print(f"{'─'*70}")
    print(f"{'SKU':<35} {'Qty':>6} {'Cost ¥':>12}")
    print(f"{'─'*70}")
    for u in updates:
        print(f"{u.sku:<35} {u.quantity_to_add:>6} {u.cost_yen_total:>12,}")
    print(f"{'─'*70}")
    print(f"{'TOTAL':<35} {total_qty:>6} {total_cost:>12,}")
    print(f"{'─'*70}")
    print(f"{len(updates)} unique SKUs, {total_qty} units, ¥{total_cost:,}")

    if unresolved:
        print(f"\n{len(unresolved)} item(s) could not be resolved to SKUs.")

    # Apply or dry-run
    if args.dry_run:
        print(f"\n[DRY RUN] No changes written.")
    else:
        store = StockStore()
        try:
            results = store.apply_updates(updates)
            print(f"\nStock updated: {len(results)} SKUs written to RDS")
        finally:
            store.close()


def cmd_import(args):
    """Seed stock store from Zoho S3 export (zoho-stock-tcg bucket)."""
    import boto3
    import openpyxl
    from io import BytesIO

    bucket = args.bucket
    keys = ['op_stock.xlsx', 'pkmn_stock.xlsx']

    s3 = boto3.client('s3')
    records = {}
    skipped_negative = 0

    for key in keys:
        print(f"Downloading s3://{bucket}/{key}...")
        obj = s3.get_object(Bucket=bucket, Key=key)
        data = obj['Body'].read()

        wb = openpyxl.load_workbook(filename=BytesIO(data), data_only=True)
        sheet = wb['Sheet1']
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = {str(h).strip().lower(): i for i, h in enumerate(header) if h}

        sku_idx = headers.get('sku')
        qty_idx = headers.get('quantity')
        if sku_idx is None or qty_idx is None:
            print(f"  Missing sku/quantity columns. Found: {list(headers.keys())}")
            continue

        count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            sku = row[sku_idx]
            qty = row[qty_idx]
            if sku is None:
                continue
            sku = str(sku).strip()
            qty = int(qty) if qty else 0

            if qty < 0:
                skipped_negative += 1
                continue
            if qty == 0:
                continue

            records[sku] = records.get(sku, 0) + qty
            count += 1

        print(f"  {key}: {count} SKUs with stock > 0")

    total_qty = sum(records.values())
    print(f"\nTotal: {len(records)} SKUs, {total_qty} units")

    if skipped_negative:
        print(f"Skipped {skipped_negative} SKUs with negative quantity")

    if args.dry_run:
        print(f"\n[DRY RUN] No changes written.")
        return

    store = StockStore()
    try:
        store.seed(records, source=f"s3://{bucket}")
        print(f"\nStock seeded: {len(records)} SKUs written to RDS")
    finally:
        store.close()


def cmd_backfill_cost(args):
    """Backfill cost for legacy stock using purchase price data."""
    store = StockStore()
    try:
        changes = store.backfill_cost(dry_run=args.dry_run)

        if not changes:
            print("No SKUs to backfill (all costs already accurate).")
            return

        prefix = "[DRY RUN] " if args.dry_run else ""
        print(f"\n{prefix}Cost backfill results:")
        print(f"{'SKU':<35} {'Seeded':>6} {'Old ¥':>10} {'New ¥':>10} {'Unit ¥':>8}")
        print(f"{'─'*73}")
        total_added = 0
        for c in sorted(changes, key=lambda x: x['sku']):
            added = c['new_cost'] - c['old_cost']
            total_added += added
            print(f"{c['sku']:<35} {c['seeded_qty']:>6} {c['old_cost']:>10,} {c['new_cost']:>10,} {c['unit_price']:>8,.0f}")
        print(f"{'─'*73}")
        print(f"{prefix}{len(changes)} SKUs backfilled, ¥{total_added:,} cost added")
    finally:
        store.close()


def cmd_sync_sales(args):
    """Pull unsynced sale events from RDS and reduce stock."""
    store = StockStore()
    try:
        conn = store._conn

        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, platform, order_id, sku, quantity, event_type "
                "FROM sales_events WHERE local_synced = FALSE "
                "ORDER BY created_at ASC"
            )
            rows = cur.fetchall()

        if not rows:
            print("No unsynced sale events.")
            return

        print(f"Found {len(rows)} unsynced event(s)\n")

        # Aggregate by SKU
        sku_totals = defaultdict(int)
        event_ids = []
        for row in rows:
            event_id, platform, order_id, sku, quantity, event_type = row
            sku_totals[sku] += quantity
            event_ids.append(event_id)

        # Build reductions (only positive net quantities are reductions)
        reductions = []
        for sku, net_qty in sorted(sku_totals.items()):
            if net_qty > 0:
                reductions.append(SaleReduction(sku=sku, quantity_sold=net_qty))

        # Show summary
        print(f"{'SKU':<35} {'Net sold':>8}")
        print(f"{'─'*45}")
        for sku, net_qty in sorted(sku_totals.items()):
            direction = 'sale' if net_qty > 0 else 'return'
            print(f"{sku:<35} {net_qty:>+8}  ({direction})")
        print(f"{'─'*45}")
        print(f"{len(sku_totals)} SKU(s), {len(reductions)} to reduce\n")

        if args.dry_run:
            if reductions:
                results = store.apply_reductions(reductions, dry_run=True)
                print("[DRY RUN] Would apply:")
                for r in results:
                    if r.get('skipped'):
                        print(f"  {r['sku']}: SKIPPED (no stock record)")
                    else:
                        clamped = ' (CLAMPED to 0)' if r['clamped'] else ''
                        print(f"  {r['sku']}: {r['old_qty']} → {r['new_qty']}{clamped}")
            print("\n[DRY RUN] No changes written.")
            return

        # Apply reductions
        if reductions:
            results = store.apply_reductions(reductions)
            print("Stock reductions applied:")
            for r in results:
                if r.get('skipped'):
                    print(f"  {r['sku']}: SKIPPED (no stock record)")
                else:
                    clamped = ' (CLAMPED to 0)' if r['clamped'] else ''
                    print(f"  {r['sku']}: {r['old_qty']} → {r['new_qty']}{clamped}")
            print(f"\n{len(results)} SKU(s) updated in RDS")

        # Mark events as local_synced
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE sales_events SET local_synced = TRUE, local_synced_at = NOW() "
                    "WHERE id = ANY(%s)",
                    (event_ids,),
                )
                conn.commit()
            print(f"{len(event_ids)} event(s) marked as local_synced")
        except Exception as e:
            print(f"Warning: Failed to mark events as synced: {e}", file=sys.stderr)
            conn.rollback()
    finally:
        store.close()


def cmd_status(args):
    """Show current stock counts."""
    store = StockStore()
    try:
        tier_config = store.get_listing_tiers()

        if args.sku:
            record = store.get(args.sku)
            if record:
                avg_cost = record.total_cost_yen / record.quantity if record.quantity > 0 else 0
                listed, reason = store.get_listed_qty_detail(args.sku)
                listed_display = str(listed) if listed is not None else "?"
                print(f"SKU:          {record.sku}")
                print(f"Actual qty:   {record.quantity}")
                print(f"Listed qty:   {listed_display}  ({reason})")
                print(f"Total cost:   ¥{record.total_cost_yen:,}")
                print(f"Avg cost:     ¥{avg_cost:,.0f}")
                print(f"Last updated: {record.last_updated}")
            else:
                print(f"No stock record for {args.sku}")
        else:
            records = store.get_all()
            if not records:
                print("No stock data. Run 'parse' to add stock from a purchase list.")
                return

            total_qty = 0
            total_listed = 0
            total_cost = 0
            unknown_count = 0
            print(f"{'SKU':<35} {'Actual':>6} {'Listed':>6} {'Cost ¥':>12} {'Avg ¥':>8}")
            print(f"{'─'*71}")
            for r in records:
                avg = r.total_cost_yen / r.quantity if r.quantity > 0 else 0
                listed = store.get_listed_qty(r.sku)
                if listed is None:
                    listed_display = "?"
                    unknown_count += 1
                else:
                    listed_display = str(listed)
                    total_listed += listed
                print(f"{r.sku:<35} {r.quantity:>6} {listed_display:>6} {r.total_cost_yen:>12,} {avg:>8,.0f}")
                total_qty += r.quantity
                total_cost += r.total_cost_yen
            print(f"{'─'*71}")
            listed_total_display = str(total_listed) if unknown_count == 0 else f"{total_listed}+?"
            print(f"{'TOTAL':<35} {total_qty:>6} {listed_total_display:>6} {total_cost:>12,}")

            if tier_config:
                tier_strs = [f"under £{t['under_gbp']}→{t['cap']}" for t in tier_config['tiers']]
                tier_strs.append(f"else→{tier_config['default_cap']}")
                print(f"\n{len(records)} SKUs in stock (tiers: {', '.join(tier_strs)})")
                if unknown_count:
                    print(f"  {unknown_count} SKU(s) missing price in cardrush_link")
            else:
                print(f"\n{len(records)} SKUs in stock (no listing tiers)")
    finally:
        store.close()


def cmd_listing(args):
    """Show, set, or clear listing tiers."""
    store = StockStore()
    try:
        if args.clear:
            store.clear_listing_tiers()
            print("Listing tiers removed (all stock listed at actual qty)")
            return

        if args.set:
            tiers = []
            default_cap = 1
            for token in args.set:
                if token.startswith('default:'):
                    default_cap = int(token.split(':')[1])
                else:
                    parts = token.split(':')
                    if len(parts) != 2:
                        print(f"Error: Invalid tier '{token}'. Use THRESHOLD:CAP (e.g. 50:4)", file=sys.stderr)
                        sys.exit(1)
                    tiers.append({"under_gbp": int(parts[0]), "cap": int(parts[1])})
            # Sort tiers by threshold ascending
            tiers.sort(key=lambda t: t['under_gbp'])
            store.set_listing_tiers(tiers, default_cap)
            _print_tiers(tiers, default_cap)
            print("\nTiers saved.")
            return

        # Default: show current tiers
        config = store.get_listing_tiers()
        if config is None:
            print("No listing tiers (all stock listed at actual qty)")
            return
        _print_tiers(config['tiers'], config['default_cap'])
    finally:
        store.close()


def _print_tiers(tiers, default_cap):
    """Print tier table."""
    print(f"{'Price range':<25} {'Max listed':>10}")
    print(f"{'─'*37}")
    for tier in tiers:
        print(f"{'Under £' + str(tier['under_gbp']):<25} {tier['cap']:>10}")
    print(f"{'£' + str(tiers[-1]['under_gbp']) + '+' if tiers else 'All':<25} {default_cap:>10}")


def cmd_export(args):
    """Export RDS stock to JSON file for offline backup."""
    store = StockStore()
    try:
        store.export_json(args.output)
        records = store.get_all()
        total_qty = sum(r.quantity for r in records)
        print(f"Exported {len(records)} SKUs ({total_qty} total qty) to {args.output}")
    finally:
        store.close()


def main():
    parser = argparse.ArgumentParser(
        prog='stock.count',
        description='Stock purchase list parser and stock count manager',
    )
    subparsers = parser.add_subparsers(dest='command', required=True)

    # parse command
    p_parse = subparsers.add_parser('parse', help='Parse a supplier purchase list')
    p_parse.add_argument('-s', '--supplier', required=True, help='Supplier name (e.g. cardrush)')
    p_parse.add_argument('-f', '--file', help='Input file (default: stdin)')
    p_parse.add_argument('--dry-run', action='store_true', help='Preview without updating stock')
    p_parse.set_defaults(func=cmd_parse)

    # import command
    p_import = subparsers.add_parser('import', help='Seed stock from Zoho S3 export')
    p_import.add_argument('--bucket', default='zoho-stock-tcg', help='S3 bucket (default: zoho-stock-tcg)')
    p_import.add_argument('--dry-run', action='store_true', help='Preview without writing')
    p_import.set_defaults(func=cmd_import)

    # backfill-cost command
    p_backfill = subparsers.add_parser('backfill-cost',
                                        help='Backfill cost for legacy stock using purchase prices')
    p_backfill.add_argument('--dry-run', action='store_true', help='Preview without writing')
    p_backfill.set_defaults(func=cmd_backfill_cost)

    # sync-sales command
    p_sync = subparsers.add_parser('sync-sales',
                                    help='Pull sale events from RDS and reduce stock')
    p_sync.add_argument('--dry-run', action='store_true', help='Preview without writing')
    p_sync.set_defaults(func=cmd_sync_sales)

    # status command
    p_status = subparsers.add_parser('status', help='Show stock counts')
    p_status.add_argument('--sku', help='Filter by specific SKU')
    p_status.set_defaults(func=cmd_status)

    # listing command
    p_listing = subparsers.add_parser('listing', help='Show or set listing tiers')
    p_listing.add_argument('--set', nargs='+', metavar='TIER',
                           help='Set tiers: THRESHOLD:CAP ... default:N (e.g. 50:4 150:2 default:1)')
    p_listing.add_argument('--clear', action='store_true', help='Remove tiers (list everything)')
    p_listing.set_defaults(func=cmd_listing)

    # export command
    p_export = subparsers.add_parser('export', help='Export RDS stock to JSON backup')
    p_export.add_argument('-o', '--output', default='stock_export.json', help='Output file path')
    p_export.set_defaults(func=cmd_export)

    args = parser.parse_args()
    args.func(args)


if __name__ == '__main__':
    main()
