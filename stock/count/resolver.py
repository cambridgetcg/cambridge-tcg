"""SKU resolution: card_number → SKU via S3 daily_prices.xlsx.

Downloads the authoritative pricing spreadsheet and builds a lookup map.
Handles standard cards and SP (special/reprint) cards.

SKU patterns:
    Standard:  OP-{card_number}-JP          e.g. OP-OP05-001-JP
    SP:        OP-{set}-SP-{card_number}-JP  e.g. OP-OP03-SP-OP01-051-JP
"""

import os
import re
from collections import defaultdict
from io import BytesIO
from typing import Dict, List, Optional, Tuple

import openpyxl

from stock.count.models import ParsedItem, ResolvedItem

S3_BUCKET = os.environ.get('S3_BUCKET', 'pricedata-tcg')
S3_KEY = os.environ.get('S3_KEY', 'daily_prices.xlsx')

# SP SKU pattern: OP-{SET}-SP-{CARD_NUMBER}-JP
SP_SKU_RE = re.compile(r'^OP-([A-Z0-9]+)-SP-([A-Z0-9-]+)-JP$')


class SKUResolver:
    """Resolves card numbers to SKUs using the S3 pricing spreadsheet."""

    def __init__(self):
        # card_number -> list of SKUs (multiple if ambiguous)
        self._card_to_skus: Dict[str, List[str]] = defaultdict(list)
        # (reprint_set, card_number) -> SKU for SP resolution
        self._sp_lookup: Dict[Tuple[str, str], str] = {}
        self._loaded = False

    def load(self, file_content: Optional[bytes] = None):
        """Load SKU mappings from S3 daily_prices.xlsx.

        Args:
            file_content: Raw xlsx bytes. If None, downloads from S3.
        """
        if file_content is None:
            import boto3
            print(f"Downloading s3://{S3_BUCKET}/{S3_KEY}...")
            s3 = boto3.client('s3')
            obj = s3.get_object(Bucket=S3_BUCKET, Key=S3_KEY)
            file_content = obj['Body'].read()

        workbook = openpyxl.load_workbook(filename=BytesIO(file_content), data_only=True)

        total = 0
        for sheet_name in ('onepiece', 'pokemon'):
            if sheet_name not in workbook.sheetnames:
                print(f"  Sheet '{sheet_name}' not found, skipping")
                continue

            sheet = workbook[sheet_name]
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = {str(cell).strip().lower(): idx for idx, cell in enumerate(header_row) if cell}

            cn_idx = headers.get('card_number')
            sku_idx = headers.get('sku')
            if cn_idx is None or sku_idx is None:
                print(f"  Sheet '{sheet_name}': missing card_number/sku columns. Found: {list(headers.keys())}")
                continue

            count = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                card_number = row[cn_idx]
                sku = row[sku_idx]
                if card_number is None or sku is None:
                    continue

                card_number = str(card_number).strip()
                sku = str(sku).strip()

                # For SP SKUs, extract (set, original_card_number) for bracket resolution
                sp_match = SP_SKU_RE.match(sku)
                if sp_match:
                    reprint_set = sp_match.group(1)
                    original_cn = sp_match.group(2)
                    self._sp_lookup[(reprint_set, original_cn)] = sku
                    # Also index by original card number for ambiguity detection
                    self._card_to_skus[original_cn].append(sku)
                else:
                    self._card_to_skus[card_number].append(sku)

                count += 1

            print(f"  {sheet_name}: {count} SKU mappings loaded")
            total += count

        self._loaded = True
        print(f"  Total: {total} SKU mappings, {len(self._sp_lookup)} SP entries")

    def resolve(self, items: List[ParsedItem]) -> List[ResolvedItem]:
        """Resolve parsed items to SKUs.

        Rules:
            - If reprint_set bracket present (e.g. [OP03]): look up SP SKU
            - If rarity is SP but no bracket: auto-search SP lookup
            - If no bracket and not SP: use standard SKU, flag ambiguous if SP variant exists
        """
        if not self._loaded:
            self.load()

        results = []
        for item in items:
            resolved = ResolvedItem(parsed=item)

            if item.reprint_set:
                # Bracket present — resolve to SP SKU
                sp_key = (item.reprint_set, item.card_number)
                sp_sku = self._sp_lookup.get(sp_key)
                if sp_sku:
                    resolved.sku = sp_sku
                    resolved.resolved = True
                else:
                    resolved.warnings.append(
                        f"SP SKU not found for {item.card_number}[{item.reprint_set}]"
                    )
            elif item.rarity and item.rarity.upper() == 'SP':
                # SP rarity without bracket — auto-search
                matching_sp = [
                    (set_name, sku)
                    for (set_name, cn), sku in self._sp_lookup.items()
                    if cn == item.card_number
                ]
                if len(matching_sp) == 1:
                    resolved.sku = matching_sp[0][1]
                    resolved.resolved = True
                elif len(matching_sp) > 1:
                    resolved.ambiguous = True
                    options = [f"{sku} (set {s})" for s, sku in matching_sp]
                    resolved.warnings.append(
                        f"Multiple SP SKUs for {item.card_number}: {', '.join(options)}. "
                        f"Use [SET] bracket to specify."
                    )
                else:
                    resolved.warnings.append(
                        f"No SP SKU found for {item.card_number}"
                    )
            else:
                # No bracket, not SP — use standard SKU
                all_skus = self._card_to_skus.get(item.card_number, [])
                standard_skus = [s for s in all_skus if '-SP-' not in s]
                sp_skus = [s for s in all_skus if '-SP-' in s]

                if standard_skus:
                    resolved.sku = standard_skus[0]
                    resolved.resolved = True

                    if sp_skus:
                        resolved.ambiguous = True
                        resolved.warnings.append(
                            f"Ambiguous: {item.card_number} has SP variants {sp_skus}. "
                            f"Defaulting to standard SKU {resolved.sku}. "
                            f"Use [SET] bracket to specify SP version."
                        )
                elif all_skus:
                    # Only SP SKUs exist — use the first and warn
                    resolved.sku = all_skus[0]
                    resolved.resolved = True
                    resolved.warnings.append(
                        f"Only SP SKU(s) found for {item.card_number}: {all_skus}"
                    )
                else:
                    resolved.warnings.append(
                        f"No SKU found for card number {item.card_number}"
                    )

            results.append(resolved)

        return results
