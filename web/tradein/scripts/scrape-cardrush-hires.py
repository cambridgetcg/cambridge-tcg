#!/usr/bin/env python3
"""
Full CardRush high-res image scraper.

Strategy:
  For each card number (e.g. OP01-001):
    1. Search cardrush-op.jp for that card number
    2. Parse all product listings returned
    3. For each listing, identify: condition, variant (base/parallel/SP), product image
    4. Pick the best (NM/A condition) image per variant
    5. Download high-res and upload to S3 under hires/{SET}/{SKU}.jpg

Output:
  - S3: jp-op-photos/hires/{SET}/{SKU}.jpg  (public-read)
  - JSON manifest: scripts/hires-manifest.json  (SKU -> S3 URL)

Usage:
  python3 scrape-cardrush-hires.py [--dry-run] [--set OP01] [--limit 10] [--workers 6]
"""

import json, os, re, sys, time, argparse, subprocess, tempfile
import urllib.request, urllib.parse, urllib.error
import concurrent.futures
from pathlib import Path

# ── Config ──────────────────────────────────────────────────────────
S3_BUCKET    = "jp-op-photos"
S3_PREFIX    = "hires"
S3_REGION    = "us-east-1"
AWS_PROFILE  = "alpha"
MANIFEST     = Path(__file__).parent / "hires-manifest.json"
BASE_URL     = "https://www.cardrush-op.jp"
SEARCH_URL   = f"{BASE_URL}/product-list/?keyword={{card_num}}&show=96"
HEADERS      = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/120",
    "Accept-Language": "ja,en-US;q=0.9",
}
DELAY        = 0.4   # per-worker request delay
IMG_RE       = re.compile(r'https://www\.cardrush-op\.jp/data/cardrush-op/product/[^\s"\'<>]+\.jpg')
CARD_NUM_RE  = re.compile(r'\{((?:OP|EB|ST|P)\d{2}-\d{3}|P-\d{3})\}')

# Condition priority (best first) — we want NM/未開封 images
# Skip: 状態B (grade B), 状態C, PSA graded (different art/holder photo), 傷あり
SKIP_CONDS   = ['状態B', '状態C', 'PSA', 'BGS', 'BVG', '傷あり', 'キズあり']
GOOD_CONDS   = ['未開封', '状態A-', '']  # blank = NM listing


def fetch(url: str) -> str | None:
    req = urllib.request.Request(url, headers=HEADERS)
    try:
        with urllib.request.urlopen(req, timeout=12) as r:
            return r.read().decode("utf-8", errors="ignore")
    except Exception:
        return None


def parse_product_page(product_id: str) -> dict | None:
    """Fetch a product page and extract: card_num, title, condition, image_url."""
    html = fetch(f"{BASE_URL}/product/{product_id}")
    if not html:
        return None

    title_m = re.search(r'<title>([^<]+)</title>', html)
    title = title_m.group(1).strip() if title_m else ""

    card_m = CARD_NUM_RE.search(title)
    card_num = card_m.group(1) if card_m else None

    img_m = IMG_RE.search(html)
    img_url = img_m.group(0) if img_m else None

    if not card_num or not img_url:
        return None

    return {
        "product_id": product_id,
        "title": title,
        "card_num": card_num,
        "img_url": img_url,
    }


def classify_variant(title: str) -> str:
    """Determine variant type from title. Returns: base | parallel | sp | alt"""
    t = title.lower()
    if 'パラレル' in title or '/p】' in t or '【l/p】' in t or '【r/p】' in t:
        return 'parallel'
    if 'sp】' in t or '【sp】' in title:
        return 'sp'
    return 'base'


def is_bad_condition(title: str) -> bool:
    return any(c in title for c in SKIP_CONDS)


def search_card(card_num: str) -> list[dict]:
    """Search CardRush for a card number, return list of product info dicts."""
    url = SEARCH_URL.format(card_num=urllib.parse.quote(card_num))
    html = fetch(url)
    if not html:
        return []

    product_ids = list(dict.fromkeys(re.findall(r'/product/(\d+)', html)))
    results = []
    for pid in product_ids:
        time.sleep(DELAY / 4)
        info = parse_product_page(pid)
        if info and info['card_num'] == card_num:
            info['variant'] = classify_variant(info['title'])
            info['bad'] = is_bad_condition(info['title'])
            results.append(info)

    return results


def best_image_per_variant(products: list[dict]) -> dict[str, str]:
    """From a list of product dicts, pick best image per variant type."""
    # Group by variant
    groups: dict[str, list] = {}
    for p in products:
        if p['bad']:
            continue
        groups.setdefault(p['variant'], []).append(p)

    best = {}
    for variant, items in groups.items():
        # Prefer: 未開封 > 状態A- > plain NM (no condition tag)
        def rank(item):
            t = item['title']
            if '未開封' in t: return 0
            if '状態A-' in t: return 1
            return 2
        items.sort(key=rank)
        best[variant] = items[0]['img_url']

    return best


def sku_from_card(card_num: str, variant: str) -> str:
    """Build SKU from card number + variant. e.g. OP01-001 + parallel -> OP-OP01-001-JP-P"""
    # card_num: OP01-001, EB01-056, ST13-011, P-105
    parts = card_num.split('-')
    if card_num.startswith('P-'):
        base = f"P-{parts[1]}-JP"
    else:
        base = f"OP-{card_num}-JP"

    if variant == 'parallel':
        return f"{base}-P"
    if variant == 'sp':
        return f"{base}-SP"
    return base


def s3_key(sku: str, set_code: str) -> str:
    return f"{S3_PREFIX}/{set_code}/{sku}.jpg"


def s3_url(key: str) -> str:
    return f"https://{S3_BUCKET}.s3.{S3_REGION}.amazonaws.com/{key}"


def already_uploaded(key: str) -> bool:
    import boto3, botocore
    if not hasattr(already_uploaded, '_s3'):
        already_uploaded._s3 = boto3.Session(profile_name=AWS_PROFILE).client('s3', region_name=S3_REGION)
    try:
        already_uploaded._s3.head_object(Bucket=S3_BUCKET, Key=key)
        return True
    except botocore.exceptions.ClientError:
        return False


def upload(img_url: str, key: str, dry_run: bool) -> bool:
    if dry_run:
        print(f"    [DRY] {img_url.split('/')[-1]} -> {key}")
        return True

    if already_uploaded(key):
        return True  # skip

    req = urllib.request.Request(img_url, headers=HEADERS)
    try:
        data = urllib.request.urlopen(req, timeout=15).read()
    except Exception as e:
        print(f"    ❌ download: {e}")
        return False

    with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as f:
        f.write(data)
        tmp = f.name

    try:
        r = subprocess.run([
            "aws", "--profile", AWS_PROFILE, "s3", "cp", tmp,
            f"s3://{S3_BUCKET}/{key}",
            "--content-type", "image/jpeg",
            "--acl", "public-read",
            "--region", S3_REGION,
        ], capture_output=True, text=True, timeout=30)
        return r.returncode == 0
    finally:
        os.unlink(tmp)


def process_card(card_num: str, dry_run: bool) -> dict:
    """Process one card number: search, find images, upload. Returns {sku: s3_url}."""
    set_code = card_num.split('-')[0] if not card_num.startswith('P-') else 'PROMO'
    time.sleep(DELAY)

    products = search_card(card_num)
    if not products:
        return {}

    best = best_image_per_variant(products)
    results = {}

    for variant, img_url in best.items():
        sku = sku_from_card(card_num, variant)
        key = s3_key(sku, set_code)
        ok = upload(img_url, key, dry_run)
        if ok:
            results[sku] = s3_url(key)

    return results


def get_all_card_numbers() -> list[str]:
    """Get full list of card numbers to scrape from price feed + S3 inventory."""
    import openpyxl

    cards = set()

    # From price feed
    try:
        wb = openpyxl.load_workbook('/tmp/pricefeed.xlsx')
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                cards.add(row[0])
    except Exception:
        pass

    # From daily_prices
    try:
        wb = openpyxl.load_workbook('/tmp/daily_prices.xlsx')
        ws = wb['onepiece']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                cards.add(row[0])
    except Exception:
        pass

    # Also infer from S3 bucket filenames (captures ST*, EB*, OP14+, PROMO)
    result = subprocess.run(
        ['aws', '--profile', AWS_PROFILE, 's3', 'ls', 's3://jp-op-photos/',
         '--recursive', '--region', S3_REGION],
        capture_output=True, text=True
    )
    for line in result.stdout.split('\n'):
        fname = line.strip().split('/')[-1] if '/' in line else ''
        # OP-OP01-001-JP.jpeg -> OP01-001
        m = re.match(r'^(?:OP|ST|EB|P)-(.+?)-JP(?:-P\d+|-CR\d+)?\.jpeg$', fname)
        if m:
            raw = m.group(1)
            # OP01-001 or ST13-011 or EB01-003
            if re.match(r'^(?:OP|EB|ST|P)\d{2}-\d{3}$', raw) or re.match(r'^P-\d{3}$', raw):
                cards.add(raw)

    return sorted(cards)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--dry-run', action='store_true')
    parser.add_argument('--set', help='Only process this set, e.g. OP01')
    parser.add_argument('--limit', type=int, default=0)
    parser.add_argument('--workers', type=int, default=6)
    args = parser.parse_args()

    cards = get_all_card_numbers()

    if args.set:
        cards = [c for c in cards if c.startswith(args.set)]

    if args.limit:
        cards = cards[:args.limit]

    print(f"{'[DRY RUN] ' if args.dry_run else ''}Scraping {len(cards)} card numbers")
    print(f"Workers: {args.workers} | Delay: {DELAY}s\n")

    # Load existing manifest
    manifest = {}
    if MANIFEST.exists():
        manifest = json.loads(MANIFEST.read_text())
    print(f"Existing manifest: {len(manifest)} entries")

    done = 0
    new_images = 0

    with concurrent.futures.ThreadPoolExecutor(max_workers=args.workers) as ex:
        futures = {ex.submit(process_card, cn, args.dry_run): cn for cn in cards}
        for future in concurrent.futures.as_completed(futures):
            cn = futures[future]
            try:
                result = future.result()
                for sku, url in result.items():
                    if sku not in manifest:
                        new_images += 1
                    manifest[sku] = url
                if result:
                    skus = ', '.join(result.keys())
                    print(f"  ✅ [{done+1}/{len(cards)}] {cn} → {len(result)} images ({skus})")
                else:
                    print(f"  ⚠️  [{done+1}/{len(cards)}] {cn} → no images found")
            except Exception as e:
                print(f"  ❌ [{done+1}/{len(cards)}] {cn} → {e}")
            done += 1

            if done % 50 == 0:
                # Save progress checkpoint
                if not args.dry_run:
                    MANIFEST.write_text(json.dumps(manifest, ensure_ascii=False, indent=2))
                print(f"\n  --- Checkpoint: {done}/{len(cards)} | manifest: {len(manifest)} ---\n")

    # Final save
    if not args.dry_run:
        MANIFEST.write_text(json.dumps(manifest, ensure_ascii=False, indent=2))

    print(f"\n{'='*60}")
    print(f"Done: {done} cards processed")
    print(f"New images: {new_images}")
    print(f"Total manifest: {len(manifest)} SKUs")
    print(f"Manifest saved: {MANIFEST}")
    print(f"{'='*60}")


if __name__ == '__main__':
    main()
